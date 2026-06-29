from __future__ import annotations

import sqlite3
import re
from io import BytesIO
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

TARGET_EMPLOYEES = ["邓晓阗", "刘柯鑫", "曹丽梅", "韩亚琳"]
SCHEDULE_DIR = Path("/Users/Zhuanz/Desktop/桌面整理/上影国际影城")
HANDOVER_DB = Path("/Users/Zhuanz/Desktop/桌面整理/上影办公/cinema-handover-assistant/data/app.db")
FENGHUANG_TOKEN_FILE = Path("/Users/Zhuanz/.hermes/workspace/fenghuang-token.txt")
DAILY_BRIEFING_TEMPLATE = Path(__file__).resolve().parent.parent / "assets" / "daily-briefing-template.png"
FONT_PATHS = [
    Path("/System/Library/Fonts/PingFang.ttc"),
    Path("/System/Library/Fonts/Supplemental/Songti.ttc"),
    Path("/System/Library/Fonts/STHeiti Light.ttc"),
]


def build_inventory_alert_lines(items: list[dict[str, Any]]) -> list[str]:
    """Return the inventory rows currently visible in the dashboard alert view."""
    lines = []
    for item in items:
        if item.get("is_excluded"):
            continue
        if item.get("status") not in {"warning", "critical"}:
            continue
        name = item.get("item_name") or "未命名商品"
        front_stock = _format_number(item.get("front_stock", 0))
        warehouse_stock = item.get("wh_stock", 0)
        warehouse_note = f" / 大仓{_format_number(warehouse_stock)}" if _number_value(warehouse_stock) < 5 else ""
        lines.append(f"• {name} → 前台{front_stock}{warehouse_note}")
    return lines


def filter_employee_performance(employees: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_name = {employee.get("name"): employee for employee in employees if employee.get("name")}
    return [by_name[name] for name in TARGET_EMPLOYEES if name in by_name]


def build_employee_lines(employees: list[dict[str, Any]]) -> list[str]:
    lines = []
    for employee in employees:
        name = employee.get("name") or "未命名"
        package_count = int(employee.get("package_count") or 0)
        activity_count = int(employee.get("activity_count") or 0)
        recharge_amount = _format_money(employee.get("recharge_amount") or 0)
        open_count = int(employee.get("open_count") or 0)
        lines.append(
            f"{name}：套餐{package_count}份 / 活动{activity_count}份 / 充值¥{recharge_amount} / 开卡{open_count}张"
        )
    return lines


def render_briefing_image_png(message: str) -> bytes:
    from PIL import Image, ImageDraw, ImageFont

    if DAILY_BRIEFING_TEMPLATE.exists():
        image = Image.open(DAILY_BRIEFING_TEMPLATE).convert("RGB")
    else:
        image = Image.new("RGB", (1672, 941), "#f4f7f2")
    draw = ImageDraw.Draw(image)
    body_font = _load_font(25)
    compact_font = _load_font(23)
    strong_font = _load_font(26, bold=True)
    label_font = _load_font(18, bold=True)
    small_font = _load_font(20)
    sections = _parse_briefing_message(message)

    overview_metrics = _build_image_overview_metrics(sections)
    shifts = sections["shift"][:4] or ["班次暂无"]
    handover_rows = [_classify_handover_image_line(line) for line in sections["handover"]]
    inventory_rows = [_classify_inventory_image_line(line) for line in sections["inventory"]]

    _draw_metric_grid(draw, (86, 236, 800, 366), overview_metrics, label_font, strong_font, body_font)
    _draw_shift_tags(draw, (858, 236, 1588, 366), shifts, label_font, compact_font)
    _draw_tagged_rows(draw, (86, 462, 802, 820), handover_rows or [("交接", "无")], label_font, body_font, small_font, max_rows=10)
    _draw_tagged_rows(draw, (858, 462, 1578, 820), inventory_rows or [("库存", "无")], label_font, body_font, small_font, max_rows=10)

    output = BytesIO()
    image.save(output, format="PNG")
    return output.getvalue()


def briefing_performance_date(target_date: date) -> date:
    return target_date - timedelta(days=1)


def format_handover_task_line(task: dict[str, Any]) -> str:
    title = str(task.get("title") or "").strip()
    detail = str(task.get("detail") or "").strip()
    task_type = str(task.get("type") or "").strip()
    if task_type == "booking" and detail:
        booking_line = _format_booking_handover_detail(detail)
        if booking_line:
            return f"• {booking_line}"
        text = detail
    elif detail and len(detail) > len(title) + 8:
        text = detail
    else:
        text = title or detail
    return f"• {text}" if text else ""


def build_briefing(repository: Any, target_date: date) -> dict[str, Any]:
    generated_on = date.today()
    yesterday = briefing_performance_date(target_date)

    schedule = read_schedule(target_date)
    cinema = read_cinema_schedule(target_date)
    prediction = read_prediction(repository, target_date)
    handover_lines = read_handover_lines(target_date)
    inventory_lines = read_inventory_lines()
    employee_lines = read_employee_lines(repository, yesterday)
    weather = read_weather(target_date)

    sections = {
        "schedule": schedule,
        "weather": weather,
        "cinema": cinema,
        "prediction": prediction,
        "handover": handover_lines,
        "inventory": inventory_lines,
        "employees": employee_lines,
        "activity": [
            "🎯 活动套票：买票送爆米花+饮料",
            "💡 前台默认出活动套票，不用询问顾客",
        ],
    }
    warnings = collect_briefing_warnings(sections)
    message = compose_message(target_date, yesterday, sections)
    return {
        "status": "ok",
        "target_date": target_date.isoformat(),
        "generated_on": generated_on.isoformat(),
        "yesterday": yesterday.isoformat(),
        "warnings": warnings,
        "sections": sections,
        "message": message,
    }


def collect_briefing_warnings(sections: dict[str, Any]) -> list[str]:
    warnings = []
    schedule = sections.get("schedule") or {}
    if not any(schedule.get(key) for key in ("early", "middle", "late", "rest")):
        warnings.append("未找到目标日期排班")
    if sections.get("weather") == "天气暂无":
        warnings.append("天气暂无")
    cinema = sections.get("cinema") or {}
    if cinema.get("first_time") == "无":
        warnings.append("影讯暂无，可能是凤凰云智 token 过期或接口暂无数据")
    if not sections.get("prediction"):
        warnings.append("预测人次暂无")
    if not sections.get("employees"):
        warnings.append("昨日业绩暂无可展示员工数据")
    return warnings


def compose_message(target_date: date, yesterday: date, sections: dict[str, Any]) -> str:
    schedule = sections["schedule"]
    cinema = sections["cinema"]
    prediction = sections["prediction"]
    handover = sections["handover"] or ["无"]
    inventory = sections["inventory"] or ["无"]
    employees = sections["employees"]

    lines = [
        f"📅 明日班次 · {target_date.month}月{target_date.day}日 {_weekday_name(target_date)}",
        f"🌤 {sections['weather']}",
        "",
        "班次安排",
        f"早班：{_join_or_none(schedule['early'])}",
        f"中班：{_join_or_none(schedule['middle'])}",
        f"晚班：{_join_or_none(schedule['late'])}",
        f"休息：{_join_or_none(schedule['rest'])}",
        "",
        "🎬 明日影讯",
        f"首场 {cinema['first_time']} 《{cinema['first_film']}》预售{cinema['first_presale']}人",
        f"全天预售 {cinema['total_presale']}人 / 预测 {prediction}人",
        "",
        "📋 待交接任务",
        *handover,
        "",
        "⚠️ 库存预警",
        *inventory,
        "",
        *sections["activity"],
    ]
    if employees:
        lines.extend(["", f"📊 昨日业绩（{yesterday.month}月{yesterday.day}日）", *employees])
    return "\n".join(lines)


def read_schedule(target_date: date) -> dict[str, list[str]]:
    try:
        import openpyxl
    except Exception:
        return _empty_schedule()

    files = sorted(SCHEDULE_DIR.glob("*排班表*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        return _empty_schedule()

    date_tokens = [f"{target_date.month}.{target_date.day}", f"{target_date.month}/{target_date.day}"]
    for file_path in files:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        try:
            for sheet in workbook.worksheets:
                target_col = _find_date_column(sheet, date_tokens)
                if not target_col:
                    continue
                header_row = target_col[0]
                return _read_schedule_sheet(sheet, header_row, target_col[1])
        finally:
            workbook.close()
    return _empty_schedule()


def read_cinema_schedule(target_date: date) -> dict[str, Any]:
    try:
        import httpx

        token = FENGHUANG_TOKEN_FILE.read_text().strip() if FENGHUANG_TOKEN_FILE.exists() else ""
        if not token:
            raise RuntimeError("missing token")
        begin = f"{target_date.isoformat()} 06:00"
        end = f"{(target_date + timedelta(days=1)).isoformat()} 05:59"
        response = httpx.post(
            f"https://lark-biprod.alibaba.com/bi/ticket/scheduleDetail?access_token={token}",
            json={"pageNo": 1, "pageSize": 200, "nullType": ["3"], "beginTime": begin, "endTime": end},
            headers={
                "Content-Type": "application/json;charset=UTF-8",
                "gray-lease-code": "16466",
                "gray-user-id": "sfcsygjxxb1",
                "Origin": "https://lark.yuekeyun.com",
            },
            timeout=30,
        )
        data = response.json()
        if data.get("code") != "SUCCESS":
            raise RuntimeError(str(data.get("message") or data.get("msg") or "request failed"))
        payload = data.get("data") or {}
        rows = payload.get("list") or []
        summary = (payload.get("summary") or {}).get("columnValueMap") or {}
        first = {}
        if rows:
            first = (sorted(rows, key=lambda row: (row.get("columnValueMap") or {}).get("startTime", ""))[0].get("columnValueMap") or {})
        return {
            "first_time": first.get("startTime") or "无",
            "first_film": first.get("filmName") or "无",
            "first_presale": int(first.get("showTicketNum") or 0),
            "total_presale": int(summary.get("showTicketNum") or 0),
        }
    except Exception:
        return {"first_time": "无", "first_film": "无", "first_presale": 0, "total_presale": 0}


def read_prediction(repository: Any, target_date: date) -> int:
    try:
        from app.services.xgboost_predictor import predict_with_xgboost

        result = predict_with_xgboost(repository, "cinema", "fenghuang", "cinema_feicuicheng", 10)
        for item in result.get("predictions", []):
            if item.get("date") == target_date.isoformat():
                return int(item.get("predicted_audience") or 0)
    except Exception:
        pass
    return 0


def read_handover_lines(target_date: date) -> list[str]:
    if not HANDOVER_DB.exists():
        return []
    query = """
        SELECT type, title, detail
        FROM tasks
        WHERE status = 'open'
          AND (
            type IN ('handover','equipment','complaint')
            OR (type = 'booking' AND due_at >= ?)
          )
        ORDER BY due_at DESC
        LIMIT 10
    """
    threshold = f"{target_date.isoformat()}T00:00:00.000Z"
    with sqlite3.connect(HANDOVER_DB) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(query, (threshold,)).fetchall()
    return [line for row in rows if (line := format_handover_task_line(dict(row)))]


def read_inventory_lines() -> list[str]:
    try:
        from app.api.routes.inventory_alert import get_all_inventory_items

        data = get_all_inventory_items(None)  # type: ignore[arg-type]
        return build_inventory_alert_lines(data.get("items", []))
    except Exception:
        return []


def read_employee_lines(repository: Any, yesterday: date) -> list[str]:
    try:
        from app.services.employee_performance import get_employee_performance

        data = get_employee_performance(repository, yesterday.isoformat(), yesterday.isoformat())
        employees = filter_employee_performance(data.get("employees", []))
        return build_employee_lines(employees)
    except Exception:
        return []


def read_weather(target_date: date) -> str:
    """从和风天气 API 获取指定日期天气预报"""
    try:
        import httpx, gzip as _gzip

        api_key = "78a66219258f4d789ca1677943dde5fb"
        api_host = "https://pe5u9xeery.re.qweatherapi.com"
        url = f"{api_host}/v7/weather/3d?location=104.08,30.67&key={api_key}"
        resp = httpx.get(url, timeout=10, headers={"Accept-Encoding": "gzip"})
        # 和风天气返回 gzip 压缩数据
        content = resp.content
        if content[:2] == b"\x1f\x8b":
            content = _gzip.decompress(content)
        data = resp.json() if content == resp.content else __import__("json").loads(content)
        if data.get("code") != "200":
            return "天气暂无"
        date_str = target_date.isoformat()
        for day in data.get("daily", []):
            if day.get("fxDate") == date_str:
                text_day = day.get("textDay", "")
                text_night = day.get("textNight", "")
                if text_day and text_night and text_day != text_night:
                    desc = f"{text_day}转{text_night}"
                else:
                    desc = text_day or text_night or "天气"
                return f"{desc} {day.get('tempMin', '?')}~{day.get('tempMax', '?')}℃"
    except Exception:
        pass
    return "天气暂无"


def _find_date_column(sheet: Any, date_tokens: list[str]) -> tuple[int, int] | None:
    for row_idx in range(1, min(sheet.max_row, 6) + 1):
        for cell in sheet[row_idx]:
            if cell.value and any(token in str(cell.value) for token in date_tokens):
                return row_idx, cell.column
    return None


def _read_schedule_sheet(sheet: Any, header_row: int, target_col: int) -> dict[str, list[str]]:
    schedule = _empty_schedule()
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        position = _cell_text(sheet, row_idx, 2) or _cell_text(sheet, row_idx, 1)
        name = _cell_text(sheet, row_idx, 3) or _cell_text(sheet, row_idx, 2)
        shift = _cell_text(sheet, row_idx, target_col) or "休"
        if not name or position == "场务":
            continue
        if shift in {"休", "假", "年假"}:
            schedule["rest"].append(name)
        elif shift == "早":
            schedule["early"].append(name)
        elif shift == "晚":
            schedule["late"].append(name)
        else:
            schedule["middle"].append(f"{name}（{shift}）")
    return schedule


def _empty_schedule() -> dict[str, list[str]]:
    return {"early": [], "middle": [], "late": [], "rest": []}


def _cell_text(sheet: Any, row: int, col: int) -> str:
    value = sheet.cell(row=row, column=col).value
    return str(value).strip() if value is not None else ""


def _join_or_none(values: list[str]) -> str:
    return "、".join(values) if values else "无"


def _weekday_name(day: date) -> str:
    return ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][day.weekday()]


def _format_number(value: Any) -> str:
    try:
        number = float(value)
        return str(int(number)) if number.is_integer() else str(number)
    except Exception:
        return str(value)


def _number_value(value: Any) -> float:
    try:
        return float(value)
    except Exception:
        return 0


def _format_money(value: Any) -> str:
    try:
        number = float(value)
        return str(int(number)) if number.is_integer() else f"{number:.2f}"
    except Exception:
        return "0"


def _parse_briefing_message(message: str) -> dict[str, Any]:
    result = {
        "title": "明日班次",
        "weather": "天气暂无",
        "shift": [],
        "film": [],
        "handover": [],
        "inventory": [],
        "activity": [],
        "performance": [],
    }
    current = ""
    for raw in message.splitlines():
        line = _image_text(raw)
        if not line:
            continue
        if line.startswith("明日班次"):
            result["title"] = line
            current = ""
        elif line.startswith("Sunny") or line.endswith("℃") or "天气暂无" in line:
            result["weather"] = line
        elif line == "班次安排":
            current = "shift"
        elif line == "明日影讯":
            current = "film"
        elif line == "待交接任务":
            current = "handover"
        elif line == "库存预警":
            current = "inventory"
        elif line.startswith("活动套票") or line.startswith("前台默认"):
            result["activity"].append(line)
        elif line.startswith("昨日业绩"):
            current = "performance"
        elif current in {"shift", "film", "handover", "inventory", "performance"}:
            result[current].append(line.lstrip("• ").strip())
    return result


def _draw_card(
    draw: Any,
    box: tuple[int, int, int, int],
    title: str,
    lines: list[str],
    title_font: Any,
    strong_font: Any,
    body_font: Any,
    small_font: Any,
    max_lines: int,
) -> None:
    x1, y1, x2, y2 = box
    draw.rounded_rectangle(box, radius=18, fill="#fbfcfa", outline="#dbe4d7", width=2)
    draw.text((x1 + 24, y1 + 34), title, fill="#145c52", font=title_font, anchor="lm")
    content_width = x2 - x1 - 48
    y = y1 + 68
    used = 0
    truncated = False
    for raw in lines:
        if used >= max_lines:
            truncated = True
            break
        if raw == "":
            y += 12
            continue
        font = strong_font if used == 0 and title in {"明日总览", "活动与业绩"} else body_font
        wrapped = _wrap_text(draw, raw, font, content_width)
        for line in wrapped:
            if used >= max_lines:
                truncated = True
                break
            draw.text((x1 + 24, y), line, fill="#1f2937", font=font)
            y += 32
            used += 1
        if truncated:
            break
    if truncated:
        draw.text((x2 - 24, y2 - 26), "更多详见文字版", fill="#64748b", font=small_font, anchor="rm")


def _draw_template_text_block(
    draw: Any,
    box: tuple[int, int, int, int],
    lines: list[str],
    strong_font: Any,
    body_font: Any,
    small_font: Any,
    max_lines: int,
    columns: int = 1,
) -> None:
    x1, y1, x2, y2 = box
    gutter = 26
    content_width = int((x2 - x1 - gutter * (columns - 1)) / columns)
    y = y1
    column = 0
    used = 0
    truncated = False
    for index, raw in enumerate(lines):
        if used >= max_lines:
            truncated = True
            break
        font = strong_font if index == 0 else body_font
        text = raw.replace("🔴", "").strip()
        wrapped = _wrap_text(draw, text, font, content_width)
        for line in wrapped:
            if used >= max_lines:
                truncated = True
                break
            if y + 32 > y2:
                if column + 1 >= columns:
                    truncated = True
                    break
                column += 1
                y = y1
            x = x1 + column * (content_width + gutter)
            draw.text((x, y), line.strip(), fill="#1f2937", font=font)
            y += 34
            used += 1
        if truncated:
            break
        y += 4
    if truncated:
        draw.text((x2, y2 - 22), "更多详见文字版", fill="#64748b", font=small_font, anchor="rm")


def _draw_metric_grid(
    draw: Any,
    box: tuple[int, int, int, int],
    metrics: list[tuple[str, str]],
    label_font: Any,
    strong_font: Any,
    body_font: Any,
) -> None:
    x1, y1, x2, _ = box
    card_height = 50
    for index, (label, value) in enumerate(metrics[:5]):
        card_width = int((x2 - x1 - 24) / 3)
        if index == 0:
            x, y, width = x1, y1, card_width
        elif index == 1:
            x, y, width = x1 + card_width + 12, y1, (x2 - x1) - card_width - 12
        else:
            x = x1 + (index - 2) * (card_width + 12)
            y = y1 + 64
            width = card_width
        draw.rounded_rectangle((x, y, x + width, y + card_height), radius=8, fill="#f7fbf7", outline="#d7e9dc", width=1)
        draw.text((x + 14, y + card_height / 2), label, fill="#2f8d59", font=label_font, anchor="lm")
        font = strong_font if index in {0, 2, 3, 4} else body_font
        label_width = 58
        display = _fit_single_line(draw, value, font, width - label_width - 14)
        draw.text((x + label_width, y + card_height / 2), display, fill="#1f2937", font=font, anchor="lm")


def _draw_shift_tags(
    draw: Any,
    box: tuple[int, int, int, int],
    lines: list[str],
    label_font: Any,
    body_font: Any,
) -> None:
    x1, y1, x2, y2 = box
    rows = lines[:4]
    gap = 12
    card_width = int((x2 - x1 - gap) / 2)
    card_height = 54
    grid_height = card_height * 2 + gap
    start_y = y1 + max(0, (y2 - y1 - grid_height) // 2)
    for index, raw in enumerate(rows):
        label, value = _split_label_value(raw)
        palette = _shift_palette(label)
        col = index % 2
        row = index // 2
        x = x1 + col * (card_width + gap)
        y = start_y + row * (card_height + gap)
        draw.rounded_rectangle((x, y, x + card_width, y + card_height), radius=8, fill=palette["card"], outline=palette["line"], width=1)
        draw.rounded_rectangle((x + 12, y + 13, x + 64, y + 41), radius=7, fill=palette["tag"])
        draw.text((x + 38, y + 27), label.replace("班", ""), fill=palette["ink"], font=label_font, anchor="mm")
        display = _fit_single_line(draw, value, body_font, card_width - 82)
        draw.text((x + 78, y + 27), display, fill="#1f2937", font=body_font, anchor="lm")


def _draw_performance_rows(
    draw: Any,
    box: tuple[int, int, int, int],
    lines: list[str],
    label_font: Any,
    body_font: Any,
    small_font: Any,
) -> None:
    x1, y1, x2, y2 = box
    if lines == ["暂无可展示数据"]:
        y = y1 + max(0, (y2 - y1 - 62) // 2)
        draw.rounded_rectangle((x1, y, x2, y + 62), radius=8, fill="#f7fbf7", outline="#d7e9dc", width=1)
        draw.text((x1 + 18, y + 31), "暂无四人业绩", fill="#64748b", font=body_font, anchor="lm")
        return
    rows = lines[:2]
    row_height = 74
    y = y1 + max(0, (y2 - y1 - row_height * len(rows)) // 2)
    for line in rows:
        name, value = _split_label_value(line)
        draw.rounded_rectangle((x1, y, x2, y + 62), radius=8, fill="#f7fbf7", outline="#d7e9dc", width=1)
        draw.text((x1 + 16, y + 31), name, fill="#176943", font=label_font, anchor="lm")
        display = _fit_single_line(draw, value, body_font, x2 - x1 - 110)
        draw.text((x1 + 104, y + 31), display, fill="#1f2937", font=body_font, anchor="lm")
        y += row_height
    if len(lines) > 2:
        draw.text((x2, y2 - 22), "更多详见文字版", fill="#64748b", font=small_font, anchor="rm")


def _draw_tagged_rows(
    draw: Any,
    box: tuple[int, int, int, int],
    rows: list[tuple[str, str]],
    label_font: Any,
    body_font: Any,
    small_font: Any,
    max_rows: int,
    columns: int = 1,
) -> None:
    x1, y1, x2, y2 = box
    gutter = 24
    col_width = int((x2 - x1 - gutter * (columns - 1)) / columns)
    row_height = 42
    per_col = max(1, (y2 - y1) // row_height)
    visible = rows[:max_rows]
    for index, (tag, text) in enumerate(visible):
        palette = _tag_palette(tag, text, index)
        column = min(columns - 1, index // per_col)
        row = index % per_col
        x = x1 + column * (col_width + gutter)
        y = y1 + row * row_height
        draw.rounded_rectangle((x, y, x + col_width, y + 36), radius=8, fill=palette["card"], outline=palette["line"], width=1)
        draw.rounded_rectangle((x + 10, y + 6, x + 58, y + 30), radius=7, fill=palette["tag"])
        draw.text((x + 34, y + 18), tag[:2], fill=palette["ink"], font=label_font, anchor="mm")
        content_width = col_width - 76
        display = _fit_single_line(draw, text, body_font, content_width)
        draw.text((x + 68, y + 18), display, fill=palette["text"], font=body_font, anchor="lm")
    if len(rows) > len(visible):
        draw.text((x2, y2 - 10), "更多详见文字版", fill="#64748b", font=small_font, anchor="rm")


def _load_font(size: int, bold: bool = False) -> Any:
    from PIL import ImageFont

    for path in FONT_PATHS:
        if path.exists():
            try:
                return ImageFont.truetype(str(path), size=size, index=1 if bold else 0)
            except Exception:
                continue
    return ImageFont.load_default()


def _text_width(draw: Any, text: str, font: Any) -> int:
    box = draw.textbbox((0, 0), text, font=font)
    return box[2] - box[0]


def _wrap_text(draw: Any, text: str, font: Any, max_width: int) -> list[str]:
    if not text:
        return [""]
    lines = []
    current = ""
    for token in _wrap_tokens(text):
        candidate = current + token
        if current and _text_width(draw, candidate, font) > max_width:
            lines.append(current)
            current = token
        else:
            current = candidate
        while _text_width(draw, current, font) > max_width and len(current) > 1:
            split_at = max(1, len(current) - 1)
            while split_at > 1 and _text_width(draw, current[:split_at], font) > max_width:
                split_at -= 1
            lines.append(current[:split_at])
            current = current[split_at:]
    if current:
        lines.append(current)
    return lines


def _wrap_tokens(text: str) -> list[str]:
    tokens: list[str] = []
    buffer = ""
    for char in text:
        if char.isascii() and (char.isalnum() or char in {" ", "-", ":", "/", "."}):
            buffer += char
            continue
        if buffer:
            tokens.append(buffer)
            buffer = ""
        tokens.append(char)
    if buffer:
        tokens.append(buffer)
    return tokens


def _fit_single_line(draw: Any, text: str, font: Any, max_width: int) -> str:
    if _text_width(draw, text, font) <= max_width:
        return text
    suffix = "…"
    current = text
    while current and _text_width(draw, current + suffix, font) > max_width:
        current = current[:-1]
    return current + suffix if current else suffix


def _split_label_value(text: str) -> tuple[str, str]:
    if "：" in text:
        label, value = text.split("：", 1)
        return label.strip(), value.strip()
    return text.strip(), ""


def _shift_palette(label: str) -> dict[str, str]:
    if label.startswith("早"):
        return {"card": "#f2fbf5", "tag": "#d9f3e3", "line": "#cbead6", "ink": "#247a4a"}
    if label.startswith("中"):
        return {"card": "#f3f8fc", "tag": "#dcecf6", "line": "#cbdfea", "ink": "#2f6f91"}
    if label.startswith("晚"):
        return {"card": "#fff8ee", "tag": "#f7e6c8", "line": "#eed7ae", "ink": "#93651f"}
    if label.startswith("休"):
        return {"card": "#f5f8f4", "tag": "#e2eadf", "line": "#d4dfd1", "ink": "#587257"}
    return {"card": "#f7fbf7", "tag": "#e8f5ec", "line": "#d7e9dc", "ink": "#247a4a"}


def _tag_palette(tag: str, text: str, index: int = 0) -> dict[str, str]:
    inventory_normal = {"card": "#f2fbf5", "tag": "#d9f3e3", "line": "#cbead6", "ink": "#247a4a", "text": "#1f2937"}
    if tag == "包场":
        return {"card": "#f2fbf5", "tag": "#d9f3e3", "line": "#cbead6", "ink": "#247a4a", "text": "#1f2937"}
    if tag == "活动":
        return {"card": "#fff8ee", "tag": "#f7e6c8", "line": "#eed7ae", "ink": "#93651f", "text": "#1f2937"}
    if tag == "客诉":
        return {"card": "#fff4f4", "tag": "#fde2e2", "line": "#f6caca", "ink": "#b42318", "text": "#1f2937"}
    if tag == "设备":
        return {"card": "#f3f8fc", "tag": "#dcecf6", "line": "#cbdfea", "ink": "#2f6f91", "text": "#1f2937"}
    if tag == "补" and "大仓" in text:
        return {"card": "#fff4f4", "tag": "#fde2e2", "line": "#f6caca", "ink": "#b42318", "text": "#7f1d1d"}
    if tag == "补":
        return inventory_normal
    return {"card": "#f7fbf7", "tag": "#e8f5ec", "line": "#d7e9dc", "ink": "#247a4a", "text": "#1f2937"}


def _compact_performance_image_line(line: str) -> str:
    return (
        line.replace("份 / 活动", " 活动")
        .replace("份 / 充值", " 充值")
        .replace(" / 开卡", " 开卡")
        .replace("¥", "")
        .replace("份", "")
        .replace("张", "")
    )


def _format_booking_handover_detail(detail: str) -> str:
    date_match = re.search(r"(\d{1,2}月\d{1,2}[日号])", detail)
    hall_match = re.search(r"(\d+\s*号厅)", detail)
    time_matches = _booking_show_time_ranges(detail)
    film_match = re.search(r"(《[^》]+》)", detail)
    if not (date_match and time_matches and film_match):
        return ""
    hall = f" {hall_match.group(1).replace(' ', '')}" if hall_match else ""
    times = " / ".join(f"{start}-{end}" for start, end in time_matches)
    return f"{date_match.group(1)}{hall} {times} {film_match.group(1)}"


def _booking_show_time_ranges(detail: str) -> list[tuple[str, str]]:
    detail = detail.replace("\\n", "\n")
    pattern = r"(\d{1,2}:\d{2})\s*[-–—]\s*(\d{1,2}:\d{2})"
    for line in detail.splitlines():
        matches = re.findall(pattern, line)
        if len(matches) > 1:
            if "映后" in line:
                return [matches[0]]
            return matches
    first_match = re.search(pattern, detail)
    return [first_match.groups()] if first_match else []


def _build_image_overview_metrics(sections: dict[str, Any]) -> list[tuple[str, str]]:
    title = str(sections.get("title") or "").replace("明日班次 ·", "").strip() or "日期暂无"
    weather = str(sections.get("weather") or "天气暂无")
    film_lines = sections.get("film") or []
    first_line = film_lines[0] if len(film_lines) > 0 else ""
    total_line = film_lines[1] if len(film_lines) > 1 else ""
    first_time = _first_match(first_line, r"首场\s+(\S+)") or "无"
    presale = _first_match(total_line, r"全天预售\s+(\d+人)") or "0人"
    prediction = _first_match(total_line, r"预测\s+(\d+人)") or "0人"
    return [
        ("日期", title),
        ("天气", weather),
        ("首场", first_time),
        ("预售", presale),
        ("预测", prediction),
    ]


def _classify_handover_image_line(line: str) -> tuple[str, str]:
    if ("号厅" in line and "《" in line) or "典映会" in line or "包场" in line:
        return "包场", line
    if "活动" in line or "套餐" in line:
        return "活动", line
    if "客诉" in line or "投诉" in line:
        return "客诉", line
    if "设备" in line or "故障" in line:
        return "设备", line
    return "交接", line


def _classify_inventory_image_line(line: str) -> tuple[str, str]:
    name, value = line, ""
    if "→" in line:
        name, value = [part.strip() for part in line.split("→", 1)]
    display_name = _short_inventory_image_name(name)
    text = f"{display_name} → {value}" if value else display_name
    return "补", text


def _short_inventory_image_name(name: str) -> str:
    if "冷饮杯" in name:
        return "冷饮杯"
    if "可乐杯" in name:
        return "可乐杯"
    return name


def _first_match(text: str, pattern: str) -> str:
    match = re.search(pattern, text)
    return match.group(1) if match else ""


def _image_text(text: str) -> str:
    for marker in ("📅", "🌤", "🎬", "📋", "⚠️", "🎯", "💡", "📊", "🔴"):
        text = text.replace(marker, "")
    return text.strip()


def _is_section_line(text: str) -> bool:
    return text in {"班次安排", "明日影讯", "待交接任务", "库存预警"} or text.startswith("昨日业绩")
