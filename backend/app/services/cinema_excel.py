import csv
import io
import json
import re
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any


PLATFORM = "fenghuang"
BUSINESS_TYPE = "cinema"
STORE_ID = "cinema_feicuicheng"
STORE_NAME = "SFC上影国际影城翡翠城店"
DISPLAY_DATA_SOURCE = "database"
DATABASE_READY_MESSAGE = "已从数据库读取凤凰云智经营数据"
DATABASE_EMPTY_MESSAGE = "暂无影院数据库快照"
DATABASE_NO_DATE_MESSAGE = "所选日期暂无影院数据"
DATABASE_SYNC_ERROR_MESSAGE = "最近同步失败"

# 娱乐项目排除列表（与 concession.py 保持一致）
_EXCLUDED_CATEGORIES = {"顽小游", "小铁台球", "顽麻社", "娱乐"}

def _is_entertainment(item: dict) -> bool:
    """判断卖品明细是否属于娱乐项目（兼容 concession_items 和 rows 两种字段格式）"""
    cat = (item.get("category") or item.get("concession_category") or "").strip()
    name = (item.get("item_name") or item.get("product_name") or item.get("concession_item_name") or "").strip()
    if cat in _EXCLUDED_CATEGORIES:
        return True
    for kw in ("顽小游", "小铁台球", "顽麻社"):
        if kw in name:
            return True
    return False

def _filtered_concession_revenue(raw: dict) -> float:
    """从原始数据中计算排除娱乐项后的卖品收入
    
    优先使用 concession_items（按订单聚合的明细），覆盖95%以上时直接过滤。
    当 concession_items 不完整时（按订单聚合可能把娱乐项合并到普通商品），
    用 rows（原始Excel行）计算娱乐项金额并从 summary 中扣除。
    """
    items = raw.get("concession_items") or []
    summary_total = raw.get("summary", {}).get("concession_revenue", 0)
    if not items:
        # 没有明细数据时，尝试用 rows 过滤
        return _fallback_filter_by_rows(raw, summary_total)
    # 计算排除项总金额（兼容 revenue 和 pay_amount 两种字段名）
    excluded_sum = sum(item.get("revenue") or item.get("pay_amount", 0) for item in items if _is_entertainment(item))
    items_total = sum(item.get("revenue") or item.get("pay_amount", 0) for item in items)
    # items 完整时（覆盖95%以上），直接用 items 过滤
    if summary_total > 0 and items_total >= summary_total * 0.95:
        return round(items_total - excluded_sum, 2)
    # items 不完整时（按订单聚合可能导致娱乐项被合并），
    # 用 rows（原始Excel行）计算娱乐项金额并从 summary 中扣除
    return _fallback_filter_by_rows(raw, summary_total)


def _fallback_filter_by_rows(raw: dict, summary_total: float) -> float:
    """用原始 rows 数据计算娱乐项金额，从 summary 中扣除"""
    rows = raw.get("rows") or []
    if not rows or summary_total <= 0:
        return summary_total
    # 用 concession_payment 字段计算娱乐项总额（rows 是原始Excel行格式）
    rows_excluded = sum(
        float(row.get("concession_payment", 0) or 0)
        for row in rows
        if _is_entertainment(row)
    )
    if rows_excluded > 0:
        return round(summary_total - rows_excluded, 2)
    return summary_total

CANONICAL_LABELS = {
    "date": "日期",
    "box_office": "票房收入",
    "customer_count": "观影人次",
    "screenings": "场次数",
    "occupancy_rate": "上座率",
    "concession_revenue": "卖品收入",
    "film_name": "影片名称",
    "film_box_office": "影片票房",
    "film_attendance": "影片人次",
}

FIELD_ALIASES = {
    "date": ["日期", "营业日期", "统计日期", "放映日期", "交易日期", "销售日期", "消费时间", "日"],
    "cinema_name": ["影院", "影院名称", "消费影院", "影城", "门店"],
    "box_office": ["票房收入", "票房", "票房总收入", "电影票房", "售票收入", "影票收入"],
    "customer_count": ["观影人次", "观影总人数", "观影人数", "购票人次", "票数"],
    "screenings": ["场次数", "场次", "放映场次", "排片场次"],
    "occupancy_rate": ["上座率", "上座率%", "平均上座率", "入座率", "满座率"],
    "concession_revenue": ["卖品收入", "卖品总收入", "卖品", "卖品销售额", "小卖收入", "商品收入"],
    "film_name": ["影片名称", "影片", "电影名称", "片名"],
    "film_box_office": ["影片票房", "影片收入", "单片票房", "电影票房", "票房（元）", "票房总金额"],
    "film_attendance": ["影片人次", "单片人次", "电影人次", "人次", "观影总人次"],
    # 卖品销售明细字段
    "concession_category": ["卖品大类"],
    "concession_sub_category": ["一级分类"],
    "concession_item_name": ["卖品名称"],
    "concession_quantity": ["销售数量"],
    "concession_original_price": ["原价（元）"],
    "concession_actual_price": ["实际售价（元）"],
    "concession_payment": ["支付金额（元）"],
    "order_no": ["订单号"],
    "operator": ["销售员", "操作员", "收银员"],
    # 会员卡消费明细字段
    "member_id": ["会员ID"],
    "card_type": ["卡类型"],
    "product_type": ["商品类型"],
    "product_name": ["商品名称"],
    "card_consume_amount": ["卡消费金额（元）"],
    # 会员开卡明细字段
    "issue_date": ["发卡日期"],
    "open_date": ["开卡日期"],
    "open_amount": ["开卡充值金额"],
    "card_number": ["卡号"],
    "card_policy": ["卡政策"],
    # 会员充值明细字段
    "recharge_date": ["充值/续费日期"],
    "recharge_amount": ["充值金额（元）"],
    "recharge_channel": ["充值渠道"],
}


class CinemaImportError(ValueError):
    pass


def parse_cinema_file(file_bytes: bytes, filename: str) -> dict[str, Any]:
    extension = Path(filename).suffix.lower()
    if extension not in {".xlsx", ".xls", ".csv"}:
        raise CinemaImportError("仅支持 .xlsx / .xls / .csv 文件")

    rows = _read_rows(file_bytes, extension)
    table_rows, matched_fields = _rows_to_records(rows)
    report_type = _detect_report_type(matched_fields, filename)
    missing_fields = _missing_fields_for_report(matched_fields, report_type)
    if not table_rows:
        raise CinemaImportError("未识别到可解析的数据行，请确认报表包含日期或经营字段")

    snapshots = _build_snapshots(table_rows, missing_fields, filename)
    if not snapshots:
        raise CinemaImportError("缺少日期字段，无法写入 daily_snapshots")

    primary = sorted(snapshots, key=lambda item: item["date"])[-1]
    return {
        "status": "ok",
        "file_name": filename,
        "report_type": report_type,
        "report_note": _report_note(report_type),
        "missing_fields": missing_fields,
        "snapshots": snapshots,
        "snapshot": primary,
        "films": primary["raw"]["films"],
        "raw": {
            "file_name": filename,
            "report_type": report_type,
            "matched_fields": sorted(matched_fields),
            "missing_fields": missing_fields,
            "rows": table_rows,
        },
    }


def save_cinema_import(repository: Any, parsed: dict[str, Any]) -> None:
    upserted_dates: set[str] = set()
    for snapshot in parsed["snapshots"]:
        snapshot = _merge_with_existing_snapshot(repository, snapshot)
        repository.upsert_daily_snapshot_values(
            business_type=BUSINESS_TYPE,
            platform=PLATFORM,
            store_id=STORE_ID,
            date=snapshot["date"],
            revenue=snapshot["revenue"],
            orders=snapshot["orders"],
            usage_rate=snapshot["usage_rate"],
            customer_count=snapshot["customer_count"],
            avg_order_value=snapshot["avg_order_value"],
            raw=snapshot["raw"],
        )
        upserted_dates.add(snapshot["date"])
    # 影片排名表：将影片数据分布到没有影片数据的日期
    if parsed["report_type"] == "film_ranking":
        films = parsed.get("films") or []
        if films:
            distribute_films_to_all_dates(repository, films, skip_dates=upserted_dates)



def _merge_with_existing_snapshot(repository: Any, snapshot: dict[str, Any]) -> dict[str, Any]:
    existing = repository.daily_snapshot_for_date(BUSINESS_TYPE, PLATFORM, STORE_ID, snapshot["date"])
    if not existing:
        return snapshot
    existing_raw = _load_raw(existing)
    incoming_raw = snapshot["raw"]
    incoming_summary = incoming_raw.get("summary", {})
    existing_summary = existing_raw.get("summary", {})
    incoming_films = incoming_raw.get("films") or []
    existing_films = existing_raw.get("films") or []
    # 合并影片列表（去重，优先用新数据）
    merged_films = incoming_films if incoming_films else existing_films
    # 合并卖品和会员数据 —— concession_detail 用 incoming 替换（明细行是原始交易记录，每行独立）
    incoming_concession = incoming_raw.get("concession_items") or []
    existing_concession = existing_raw.get("concession_items") or []
    incoming_member = incoming_raw.get("member_items") or []
    existing_member = existing_raw.get("member_items") or []
    incoming_recharge = incoming_raw.get("member_recharge_items") or []
    existing_recharge = existing_raw.get("member_recharge_items") or []
    incoming_open_card = incoming_raw.get("member_open_card_items") or []
    existing_open_card = existing_raw.get("member_open_card_items") or []
    # 卖品：有新数据时直接替换（明细表每行是独立交易，不能按品名去重）
    merged_concession = incoming_concession if incoming_concession else existing_concession
    # 会员消费：有新数据时直接替换
    merged_member = incoming_member if incoming_member else existing_member
    # 会员充值：有新数据时直接替换
    merged_recharge = incoming_recharge if incoming_recharge else existing_recharge
    # 会员开卡：有新数据时直接替换
    merged_open_card = incoming_open_card if incoming_open_card else existing_open_card
    # 影片排名表不覆盖已有营运数据（避免累计票房写成单日值）
    incoming_report_type = incoming_raw.get("report_type", "")
    existing_has_operations = existing_summary.get("box_office", 0) > 0
    if incoming_report_type == "film_ranking" and existing_has_operations:
        # 保留已有 summary，只合并影片列表
        best_box_office = existing_summary.get("box_office", 0)
        best_concession = existing_summary.get("concession_revenue", 0)
        best_customer = existing_summary.get("customer_count", 0)
        best_screenings = existing_summary.get("screenings", 0)
        # 上座率：取非零值（场次放映明细有上座率数据）
        best_occupancy = _best_value(existing_summary.get("occupancy_rate"), incoming_summary.get("occupancy_rate"))
        best_member_consume = existing_summary.get("member_consume", 0)
    else:
        # 取各字段最优值（非零优先）
        best_box_office = _best_value(existing_summary.get("box_office"), incoming_summary.get("box_office"))
        best_concession = _best_value(existing_summary.get("concession_revenue"), incoming_summary.get("concession_revenue"))
        best_customer = _best_int(existing_summary.get("customer_count"), incoming_summary.get("customer_count"))
        best_screenings = _best_int(existing_summary.get("screenings"), incoming_summary.get("screenings"))
        best_occupancy = _best_value(existing_summary.get("occupancy_rate"), incoming_summary.get("occupancy_rate"))
        best_member_consume = _best_value(existing_summary.get("member_consume"), incoming_summary.get("member_consume"))
    revenue = round(best_box_office + best_concession, 2)
    avg_order_value = round(revenue / best_customer, 2) if best_customer else 0
    merged_raw = {
        **existing_raw,
        **incoming_raw,
        "films": merged_films,
        "concession_items": merged_concession,
        "member_items": merged_member,
        "member_recharge_items": merged_recharge,
        "member_open_card_items": merged_open_card,
        "summary": {
            "box_office": best_box_office,
            "concession_revenue": best_concession,
            "customer_count": best_customer,
            "screenings": best_screenings,
            "occupancy_rate": best_occupancy,
            "member_consume": best_member_consume,
            "member_recharge_total": sum(item.get("amount", 0) for item in merged_recharge),
            "member_open_card_total": sum(item.get("amount", 0) for item in merged_open_card),
            "revenue": revenue,
            "avg_order_value": avg_order_value,
        },
        "imported_reports": list(set([
            *(existing_raw.get("imported_reports") or [existing_raw.get("file_name")]),
            incoming_raw.get("file_name"),
        ])),
        "missing_fields": [
            f for f in set(
                existing_raw.get("missing_fields", []) + incoming_raw.get("missing_fields", [])
            )
            if f not in _fields_covered_by_report(incoming_raw.get("file_name", ""))
        ],
    }
    return {
        "date": snapshot["date"],
        "revenue": revenue,
        "box_office": best_box_office,
        "concession_revenue": best_concession,
        "customer_count": best_customer,
        "orders": best_screenings,
        "usage_rate": best_occupancy,
        "avg_order_value": avg_order_value,
        "raw": merged_raw,
    }


def distribute_films_to_all_dates(repository: Any, films: list[dict[str, Any]], skip_dates: set[str] | None = None) -> None:
    """将影片数据分布到所有已有的影院快照中（不覆盖已有影片数据的快照）"""
    if not films:
        return
    snapshots = repository.daily_snapshots_for(BUSINESS_TYPE, PLATFORM, STORE_ID, 90)
    for snapshot in snapshots:
        if skip_dates and snapshot["date"] in skip_dates:
            continue
        raw = _load_raw(snapshot)
        # 跳过已有影片数据的快照（营运报表已有正确的单日影片数据）
        if raw.get("films"):
            continue
        raw["films"] = films
        repository.upsert_daily_snapshot_values(
            business_type=BUSINESS_TYPE,
            platform=PLATFORM,
            store_id=STORE_ID,
            date=snapshot["date"],
            revenue=snapshot["revenue"],
            orders=snapshot["orders"],
            usage_rate=snapshot["usage_rate"],
            customer_count=snapshot["customer_count"],
            avg_order_value=snapshot["avg_order_value"],
            raw=raw,
        )


def _best_value(existing: Any, incoming: Any) -> float:
    """取非零值优先"""
    e = float(existing or 0)
    i = float(incoming or 0)
    return i if i > 0 else e


def _best_int(existing: Any, incoming: Any) -> int:
    """取非零整数优先"""
    e = int(existing or 0)
    i = int(incoming or 0)
    return i if i > 0 else e


def _fields_covered_by_report(filename: str) -> list[str]:
    """根据文件名判断该报表覆盖哪些字段"""
    if "营运" in filename:
        return ["日期", "票房收入", "观影人次", "场次数", "上座率", "卖品收入"]
    if "影片" in filename:
        return ["影片名称", "影片票房", "影片人次"]
    if "卖品" in filename:
        return ["卖品收入"]
    if "会员" in filename:
        return ["会员消费"]
    return []


def cinema_overview(repository: Any, target_date: str | None = None, days: int = 1, start_date: str | None = None) -> dict[str, Any]:
    snapshot = _selected_snapshot(repository, target_date)
    latest_log = repository.latest_sync_log_for_platform(PLATFORM)
    if latest_log and latest_log.get("status") == "failed" and not snapshot:
        return _empty_overview(
            "error",
            DATABASE_SYNC_ERROR_MESSAGE,
            target_date=target_date,
            last_import_time=latest_log.get("finished_at") or latest_log.get("started_at"),
        )
    if not snapshot:
        has_any_snapshot = repository.latest_daily_snapshot_for(BUSINESS_TYPE, PLATFORM, STORE_ID)
        if target_date and has_any_snapshot:
            return _empty_overview("no_data", DATABASE_NO_DATE_MESSAGE, target_date=target_date)
        return _empty_overview("not_imported", DATABASE_EMPTY_MESSAGE, target_date=target_date)

    # 范围模式：聚合数据（支持 start_date 或 days）
    if start_date or days > 1:
        max_date = snapshot["date"]
        snapshots = repository.daily_snapshots_for(BUSINESS_TYPE, PLATFORM, STORE_ID, days, max_date=max_date, start_date=start_date)
        total_box = 0
        total_concession = 0
        total_customer = 0
        total_screenings = 0
        for s in snapshots:
            r = _load_raw(s)
            summary = r.get("summary", {})
            total_box += summary.get("box_office", 0)
            total_concession += _filtered_concession_revenue(r)
            total_customer += s.get("customer_count", 0) or 0
            total_screenings += s.get("orders", 0) or 0
        total_revenue = round(total_box + total_concession, 2)
        avg_order = round(total_box / total_customer, 2) if total_customer else 0
        latest_import_time = _latest_success_time(repository)
        return {
            "status": "ok",
            "data_source": DISPLAY_DATA_SOURCE,
            "date": f"{snapshots[0]['date']}~{snapshots[-1]['date']}" if snapshots else snapshot["date"],
            "revenue": total_revenue,
            "box_office": round(total_box, 2),
            "concession_revenue": round(total_concession, 2),
            "customer_count": total_customer,
            "screenings": total_screenings,
            "occupancy_rate": 0,
            "avg_order_value": avg_order,
            "last_import_time": latest_import_time or snapshot["created_at"],
            "message": DATABASE_READY_MESSAGE,
        }

    raw = _load_raw(snapshot)
    latest_import_time = _latest_success_time(repository)
    filtered_concession = _filtered_concession_revenue(raw)
    box_office = raw.get("summary", {}).get("box_office", 0)
    customers = snapshot["customer_count"] or 0
    return {
        "status": "ok",
        "data_source": DISPLAY_DATA_SOURCE,
        "date": snapshot["date"],
        "revenue": round(box_office + filtered_concession, 2),
        "box_office": box_office,
        "concession_revenue": round(filtered_concession, 2),
        "customer_count": customers,
        "screenings": raw.get("summary", {}).get("screenings", snapshot["orders"]),
        "occupancy_rate": snapshot["usage_rate"],
        "avg_order_value": round(box_office / customers, 2) if customers else 0,
        "last_import_time": latest_import_time or snapshot["created_at"],
        "message": DATABASE_READY_MESSAGE,
    }


def cinema_detail(repository: Any, target_date: str | None = None, days: int = 30, start_date: str | None = None) -> dict[str, Any]:
    today = date.today().isoformat()
    snapshot = _selected_snapshot(repository, target_date)
    if not snapshot:
        has_any_snapshot = repository.latest_daily_snapshot_for(BUSINESS_TYPE, PLATFORM, STORE_ID)
        if target_date and has_any_snapshot:
            return {
                "status": "no_data",
                "data_source": DISPLAY_DATA_SOURCE,
                "date": target_date,
                "message": DATABASE_NO_DATE_MESSAGE,
            }
        return {
            "status": "not_imported",
            "data_source": DISPLAY_DATA_SOURCE,
            "message": DATABASE_EMPTY_MESSAGE,
        }

    max_date = snapshot["date"] if target_date else today
    snapshots_30d = repository.daily_snapshots_for(BUSINESS_TYPE, PLATFORM, STORE_ID, days, max_date=max_date, start_date=start_date)
    trend_30d = [_snapshot_trend_item(item) for item in snapshots_30d]
    trend_7d = trend_30d[-7:]
    latest_raw = _load_raw(snapshot)
    imports = repository.latest_sync_logs(platform=PLATFORM, limit=10)

    # 范围模式：today 字段用聚合数据（支持 start_date 或 days > 1）
    is_range_mode = (start_date or days > 1) and not target_date and len(snapshots_30d) > 1
    if is_range_mode:
        total_box = 0
        total_concession = 0
        total_customer = 0
        total_screenings = 0
        for s in snapshots_30d:
            r = _load_raw(s)
            summary = r.get("summary", {})
            total_box += summary.get("box_office", 0)
            total_concession += _filtered_concession_revenue(r)
            total_customer += s.get("customer_count", 0) or 0
            total_screenings += s.get("orders", 0) or 0
        total_revenue = round(total_box + total_concession, 2)
        avg_order = round(total_box / total_customer, 2) if total_customer else 0
        today_data = {
            "date": f"{snapshots_30d[-1]['date']}~{snapshots_30d[0]['date']}",
            "box_office": round(total_box, 2),
            "concession_revenue": round(total_concession, 2),
            "customer_count": total_customer,
            "screenings": total_screenings,
            "occupancy_rate": 0,
            "revenue": total_revenue,
            "avg_order_value": avg_order,
        }
    else:
        today_data = _overview_from_snapshot(snapshot, latest_raw)

    # 影片数据：按影片名聚合，避免同一影片多版本/多场次拆成多行
    if is_range_mode:
        films = _aggregate_films(snapshots_30d)
    else:
        films = _aggregate_film_items(latest_raw.get("films", []))
        if not films:
            for snap in reversed(snapshots_30d):
                raw = _load_raw(snap)
                if raw.get("films"):
                    films = _aggregate_film_items(raw["films"])
                    break

    return {
        "status": "ok",
        "data_source": DISPLAY_DATA_SOURCE,
        "today": today_data,
        "box_office_trend_7d": trend_7d,
        "box_office_trend_30d": trend_30d,
        "film_box_office_ranking": sorted(films, key=lambda item: item["film_box_office"], reverse=True),
        "film_attendance_ranking": sorted(films, key=lambda item: item["film_attendance"], reverse=True),
        "screening_analysis": [
            {
                "date": item["date"],
                "screenings": item["orders"],
                "occupancy_rate": item["usage_rate"],
            }
            for item in snapshots_30d
        ],
        "recent_imports": [
            {
                "file_name": item.get("file_name"),
                "import_time": item.get("finished_at") or item.get("started_at"),
                "status": item.get("status"),
                "error_reason": item.get("message") if item.get("status") == "failed" else None,
                "message": item.get("message"),
            }
            for item in imports
        ],
        "missing_fields": latest_raw.get("missing_fields", []),
        "message": DATABASE_READY_MESSAGE,
    }


def _normalize_film_metrics(films: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for film in films:
        name = film.get("film_name") or film.get("name")
        if not name:
            continue
        normalized.append(
            {
                **film,
                "film_name": name,
                "film_box_office": film.get("film_box_office", film.get("box_office", 0)) or 0,
                "film_attendance": film.get("film_attendance", film.get("audience", 0)) or 0,
            }
        )
    return normalized


def _aggregate_film_items(films: list[dict[str, Any]]) -> list[dict[str, Any]]:
    totals: dict[str, dict[str, Any]] = {}
    for film in _normalize_film_metrics(films):
        name = film["film_name"]
        item = totals.setdefault(name, {"film_name": name, "film_box_office": 0.0, "film_attendance": 0})
        item["film_box_office"] = round(item["film_box_office"] + float(film.get("film_box_office") or 0), 2)
        item["film_attendance"] += int(film.get("film_attendance") or 0)
    return list(totals.values())


def cinema_status(repository: Any) -> dict[str, Any]:
    snapshot = repository.latest_daily_snapshot_for(BUSINESS_TYPE, PLATFORM, STORE_ID, max_date=date.today().isoformat())
    latest_log = repository.latest_sync_log_for_platform(PLATFORM)
    if latest_log and latest_log.get("status") == "failed" and (
        not snapshot or (latest_log.get("started_at") or "") >= (snapshot.get("created_at") or "")
    ):
        return {
            "platform": PLATFORM,
            "business_type": BUSINESS_TYPE,
            "status": "error",
            "data_source": DISPLAY_DATA_SOURCE,
            "last_sync_time": latest_log.get("finished_at") or latest_log.get("started_at"),
            "message": DATABASE_SYNC_ERROR_MESSAGE,
            "error_reason": latest_log.get("message"),
        }
    if snapshot:
        return {
            "platform": PLATFORM,
            "business_type": BUSINESS_TYPE,
            "status": "ok",
            "data_source": DISPLAY_DATA_SOURCE,
            "last_sync_time": _latest_success_time(repository) or snapshot["created_at"],
            "message": DATABASE_READY_MESSAGE,
        }
    return {
        "platform": PLATFORM,
        "business_type": BUSINESS_TYPE,
        "status": "not_imported",
        "data_source": DISPLAY_DATA_SOURCE,
        "last_sync_time": None,
        "message": DATABASE_EMPTY_MESSAGE,
    }


def _read_rows(file_bytes: bytes, extension: str) -> list[list[Any]]:
    if extension == ".csv":
        text = file_bytes.decode("utf-8-sig")
        return [row for row in csv.reader(io.StringIO(text))]
    if extension == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        if sheet.max_row == 1 and sheet.max_column == 1:
            sheet.reset_dimensions()
        rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        # openpyxl read_only 模式有时因维度检测问题只读到1行，用 pandas 重试
        if len(rows) <= 1:
            try:
                import pandas as pd
                df = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=object)
                df = df.fillna("")
                rows = [list(row) for row in df.itertuples(index=False, name=None)]
            except Exception:
                pass  # pandas 失败就保留 openpyxl 结果
        return rows
    if extension == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(file_contents=file_bytes)
        sheet = workbook.sheet_by_index(0)
        return [[sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)] for row_index in range(sheet.nrows)]
    raise CinemaImportError("仅支持 .xlsx / .xls / .csv 文件")


def _rows_to_records(rows: list[list[Any]]) -> tuple[list[dict[str, Any]], set[str]]:
    header_index, header_map = _find_header(rows)
    if header_index is None:
        return [], set()

    records: list[dict[str, Any]] = []
    matched_fields = set(header_map.values())
    for row in rows[header_index + 1:]:
        if not any(_clean_cell(value) for value in row):
            continue
        # 跳过合计行（所有字段都是 '--'）和元数据行（影院名称: / 下载人:）
        first_cell = _clean_cell(row[0]) if row else ""
        if first_cell == "--":
            continue
        if first_cell.startswith("影院名称:") or first_cell.startswith("下载人:"):
            continue
        record: dict[str, Any] = {}
        for index, field in header_map.items():
            record[field] = row[index] if index < len(row) else None
        # 跳过含 '--' 值的合计行（有些文件只在部分列有 '--'）
        if all(_clean_cell(v) in ("", "--") for v in record.values()):
            continue
        if any(_clean_cell(value) for value in record.values()):
            records.append(record)
    filtered = _filter_store_rows(records)
    if filtered:
        records = filtered
    return records, matched_fields


def _find_header(rows: list[list[Any]]) -> tuple[int | None, dict[int, str]]:
    best_index: int | None = None
    best_map: dict[int, str] = {}
    for index, row in enumerate(rows[:30]):
        current: dict[int, str] = {}
        claimed: set[str] = set()
        for column_index, value in enumerate(row):
            normalized = _normalize_header(value)
            if not normalized:
                continue
            for field, aliases in FIELD_ALIASES.items():
                if field in claimed:
                    continue
                if normalized in {_normalize_header(alias) for alias in aliases}:
                    current[column_index] = field
                    claimed.add(field)
                    break
        if len(current) > len(best_map):
            best_index = index
            best_map = current
    return best_index, best_map


def _detect_report_type(matched_fields: set[str], filename: str) -> str:
    name = filename.lower()
    if "影片成绩" in filename:
        return "film_ranking"
    if "卖品销售" in filename or "concession_item_name" in matched_fields or "concession_category" in matched_fields:
        return "concession_detail"
    if "会员卡" in filename and "充值" in filename:
        return "member_recharge"
    if "会员卡" in filename and "开卡" in filename:
        return "member_open_card"
    if "会员卡" in filename or "member_id" in matched_fields or "card_consume_amount" in matched_fields:
        return "member_detail"
    if {"recharge_date", "recharge_amount"}.intersection(matched_fields):
        return "member_recharge"
    if {"issue_date", "open_amount"}.intersection(matched_fields) or {"open_date", "open_amount"}.intersection(matched_fields):
        return "member_open_card"
    if {"date", "box_office", "customer_count", "screenings"}.intersection(matched_fields):
        return "operations"
    if {"film_name", "film_box_office", "film_attendance"}.issubset(matched_fields):
        return "film_ranking"
    if "film" in name:
        return "film_ranking"
    return "generic"


def _missing_fields_for_report(matched_fields: set[str], report_type: str) -> list[str]:
    required_by_type = {
        "operations": ["date", "box_office", "customer_count", "screenings", "occupancy_rate", "concession_revenue"],
        "film_ranking": ["film_name", "film_box_office", "film_attendance"],
        "concession_detail": ["date", "concession_revenue"],
        "member_detail": ["date", "customer_count"],
        "member_recharge": ["recharge_date", "recharge_amount"],
        "member_open_card": ["issue_date", "open_amount"],
        "generic": list(CANONICAL_LABELS),
    }
    required = required_by_type.get(report_type, required_by_type["generic"])
    return [CANONICAL_LABELS[field] for field in required if field in CANONICAL_LABELS and field not in matched_fields]


def _report_note(report_type: str) -> str:
    return {
        "operations": "营运综合报表用于写入每日票房、人次、场次、卖品等核心经营数据。",
        "film_ranking": "影片成绩排名表用于补充影片票房和人次排行，不覆盖同日营运综合数据。",
        "concession_detail": "卖品销售明细用于补充卖品收入明细。",
        "member_detail": "会员卡明细用于补充会员消费信息。",
        "member_recharge": "会员充值明细用于记录会员充值金额和笔数。",
        "member_open_card": "会员开卡明细用于记录新会员开卡数量和金额。",
    }.get(report_type, "已按通用表格解析。")


def _build_snapshots(rows: list[dict[str, Any]], missing_fields: list[str], filename: str) -> list[dict[str, Any]]:
    report_type = _detect_report_type(set(rows[0].keys()) if rows else {}, filename)
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    fallback_date = _date_range_end_from_filename(filename)
    for row in rows:
        # 根据报表类型选择日期字段
        if report_type == "member_recharge":
            parsed_date = _parse_date(row.get("recharge_date"))
        elif report_type == "member_open_card":
            parsed_date = _parse_date(row.get("issue_date")) or _parse_date(row.get("open_date"))
        else:
            parsed_date = _parse_date(row.get("date"))
        if not parsed_date:
            parsed_date = fallback_date
        if parsed_date:
            grouped[parsed_date].append(row)
    snapshots: list[dict[str, Any]] = []
    # 影片排名表：汇总所有行的影片数据，分布到日期范围内的每一天
    all_films = []
    if report_type == "film_ranking":
        for date_rows in grouped.values():
            for row in date_rows:
                film = _film_from_row(row)
                if film:
                    all_films.append(film)
        # 按影片名聚合
        film_totals: dict[str, dict[str, Any]] = {}
        for film in all_films:
            name = film["film_name"]
            if name not in film_totals:
                film_totals[name] = {"film_name": name, "film_box_office": 0, "film_attendance": 0}
            film_totals[name]["film_box_office"] += film["film_box_office"]
            film_totals[name]["film_attendance"] += film["film_attendance"]
        all_films = list(film_totals.values())
    for snapshot_date, date_rows in grouped.items():
        raw_films = [_film_from_row(row) for row in date_rows if _film_from_row(row)]
        # 按影片名聚合（同一影片多场次合并）
        film_agg: dict[str, dict[str, Any]] = {}
        for f in raw_films:
            name = f["film_name"]
            if name not in film_agg:
                film_agg[name] = {"film_name": name, "film_box_office": 0.0, "film_attendance": 0}
            film_agg[name]["film_box_office"] = round(film_agg[name]["film_box_office"] + f["film_box_office"], 2)
            film_agg[name]["film_attendance"] += f["film_attendance"]
        films = list(film_agg.values())
        has_explicit_row_date = any(_parse_date(row.get("date")) for row in date_rows)
        box_office = sum(item["film_box_office"] for item in films) if films else _first_number(date_rows, "box_office")
        if not box_office and "票房收入" in missing_fields:
            box_office = sum(item["film_box_office"] for item in films)
        concession_revenue = _sum_or_first(date_rows, "concession_revenue", prefer_sum=not bool(films))
        customer_count = int(round(sum(item["film_attendance"] for item in films))) if films else int(round(_first_number(date_rows, "customer_count")))
        if not customer_count and "观影人次" in missing_fields:
            customer_count = int(round(sum(item["film_attendance"] for item in films)))
        screenings = int(round(_sum_or_first(date_rows, "screenings", prefer_sum=(report_type == "film_ranking" or not bool(films)))))
        occupancy_rate = _first_rate(date_rows, "occupancy_rate")
        # 如果没有影片票房数据但有 film_name + screenings，按场次比例拆分总票房/人次
        if not films and box_office:
            film_screenings: dict[str, int] = {}
            for row in date_rows:
                fname = _clean_cell(row.get("film_name"))
                if not fname:
                    continue
                fs = int(round(_parse_number(row.get("screenings")) or 0))
                if fs > 0:
                    film_screenings[fname] = film_screenings.get(fname, 0) + fs
            total_fs = sum(film_screenings.values())
            if total_fs > 0:
                for fname, fs in film_screenings.items():
                    ratio = fs / total_fs
                    films.append({
                        "film_name": fname,
                        "film_box_office": round(box_office * ratio, 2),
                        "film_attendance": int(round(customer_count * ratio)),
                    })
        # 卖品明细数据
        concession_items = []
        if report_type == "concession_detail":
            for row in date_rows:
                item_name = _clean_cell(row.get("concession_item_name"))
                if not item_name:
                    continue
                category = _clean_cell(row.get("concession_category"))
                sub_category = _clean_cell(row.get("concession_sub_category"))
                quantity = int(_parse_number(row.get("concession_quantity")))
                actual_price = _parse_number(row.get("concession_actual_price"))
                operator = _clean_cell(row.get("operator"))
                payment_value = row.get("concession_payment")
                revenue = _parse_number(payment_value)
                if payment_value in (None, ""):
                    revenue = actual_price

                if revenue != 0 or quantity != 0:
                    concession_items.append({
                        "item_name": item_name,
                        "category": category,
                        "sub_category": sub_category,
                        "quantity": quantity,
                        "revenue": revenue,
                        "operator": operator,
                    })
            if concession_items and not concession_revenue:
                concession_revenue = sum(item["revenue"] for item in concession_items)
        # 会员消费数据
        member_items = []
        member_consume = 0
        if report_type == "member_detail":
            for row in date_rows:
                member_id = _clean_cell(row.get("member_id"))
                product_type = _clean_cell(row.get("product_type"))
                product_name = _clean_cell(row.get("product_name"))
                card_amount = _parse_number(row.get("card_consume_amount"))
                operator = _clean_cell(row.get("operator"))
                if card_amount > 0:
                    member_items.append({
                        "member_id": member_id,
                        "product_type": product_type,
                        "product_name": product_name,
                        "amount": card_amount,
                        "operator": operator,
                    })
            member_consume = sum(item["amount"] for item in member_items)
            # 会员表不设置 customer_count，保留已有数据
            customer_count = 0
        # 会员充值数据
        member_recharge_items = []
        member_recharge_total = 0
        if report_type == "member_recharge":
            for row in date_rows:
                member_id = _clean_cell(row.get("member_id"))
                card_number = _clean_cell(row.get("card_number"))
                card_type = _clean_cell(row.get("card_type"))
                recharge_amount = _parse_number(row.get("recharge_amount"))
                operator = _clean_cell(row.get("operator"))
                if recharge_amount > 0:
                    member_recharge_items.append({
                        "member_id": member_id,
                        "card_number": card_number,
                        "card_type": card_type,
                        "amount": recharge_amount,
                        "operator": operator,
                    })
            member_recharge_total = sum(item["amount"] for item in member_recharge_items)
            # 充值表不设置 customer_count 和 box_office，保留已有数据
            customer_count = 0
            box_office = 0
        # 会员开卡数据
        member_open_card_items = []
        member_open_card_total = 0
        if report_type == "member_open_card":
            for row in date_rows:
                member_id = _clean_cell(row.get("member_id"))
                card_number = _clean_cell(row.get("card_number"))
                card_type = _clean_cell(row.get("card_type"))
                open_amount = _parse_number(row.get("open_amount"))
                operator = _clean_cell(row.get("operator"))
                if open_amount > 0:
                    member_open_card_items.append({
                        "member_id": member_id,
                        "card_number": card_number,
                        "card_type": card_type,
                        "amount": open_amount,
                        "operator": operator,
                    })
            member_open_card_total = sum(item["amount"] for item in member_open_card_items)
            # 开卡表不设置 customer_count 和 box_office，保留已有数据
            customer_count = 0
            box_office = 0
        # 影片排名表：仅当只有1个日期（累计报表）时才用汇总数据；
        # 多日期文件（如场次放映明细）保留每日独立影片数据
        if report_type == "film_ranking" and all_films and len(grouped) == 1:
            films = all_films
        revenue = round(box_office + concession_revenue, 2)
        avg_order_value = round(revenue / customer_count, 2) if customer_count else 0
        raw = {
            "file_name": filename,
            "report_type": report_type,
            "date": snapshot_date,
            "store_name": STORE_NAME,
            "missing_fields": missing_fields,
            "summary": {
                "box_office": box_office,
                "concession_revenue": concession_revenue,
                "customer_count": customer_count,
                "screenings": screenings,
                "occupancy_rate": occupancy_rate,
                "member_consume": member_consume,
                "member_recharge_total": member_recharge_total,
                "member_open_card_total": member_open_card_total,
                "revenue": revenue,
                "avg_order_value": avg_order_value,
            },
            "films": films,
            "concession_items": concession_items,
            "member_items": member_items,
            "member_recharge_items": member_recharge_items,
            "member_open_card_items": member_open_card_items,
            "rows": date_rows,
        }
        snapshots.append(
            {
                "date": snapshot_date,
                "revenue": revenue,
                "box_office": box_office,
                "concession_revenue": concession_revenue,
                "customer_count": customer_count,
                "orders": screenings,
                "usage_rate": occupancy_rate,
                "avg_order_value": avg_order_value,
                "raw": raw,
            }
        )
    return snapshots


def _film_from_row(row: dict[str, Any]) -> dict[str, Any] | None:
    film_name = _clean_cell(row.get("film_name"))
    if not film_name:
        return None
    film_box_office = _parse_number(row.get("film_box_office")) or _parse_number(row.get("box_office"))
    film_attendance = int(round(_parse_number(row.get("film_attendance")) or _parse_number(row.get("customer_count"))))
    if not film_box_office and not film_attendance:
        return None
    return {
        "film_name": film_name,
        "film_box_office": film_box_office,
        "film_attendance": film_attendance,
    }


def _aggregate_films(snapshots: list[dict[str, Any]]) -> list[dict[str, Any]]:
    totals: dict[str, dict[str, Any]] = {}
    for snapshot in snapshots:
        raw = _load_raw(snapshot)
        for film in _normalize_film_metrics(raw.get("films", [])):
            name = film.get("film_name") or "未命名影片"
            item = totals.setdefault(name, {"film_name": name, "film_box_office": 0.0, "film_attendance": 0})
            item["film_box_office"] = round(item["film_box_office"] + float(film.get("film_box_office") or 0), 2)
            item["film_attendance"] += int(film.get("film_attendance") or 0)
    return list(totals.values())


def _snapshot_trend_item(snapshot: dict[str, Any]) -> dict[str, Any]:
    raw = _load_raw(snapshot)
    return {
        "date": snapshot["date"],
        "box_office": raw.get("summary", {}).get("box_office", 0),
        "revenue": snapshot["revenue"],
        "customer_count": snapshot["customer_count"],
        "screenings": snapshot["orders"],
        "occupancy_rate": snapshot["usage_rate"],
    }


def _sync_raw_summary(raw: dict[str, Any], snapshot: dict[str, Any]) -> dict[str, Any]:
    summary = raw.get("summary", {})
    raw["summary"] = {
        **summary,
        "box_office": summary.get("box_office", snapshot["revenue"]),
        "concession_revenue": summary.get("concession_revenue", 0),
        "customer_count": snapshot["customer_count"],
        "screenings": snapshot["orders"],
        "occupancy_rate": snapshot["usage_rate"],
        "revenue": snapshot["revenue"],
        "avg_order_value": snapshot["avg_order_value"],
    }
    return raw


def _overview_from_snapshot(snapshot: dict[str, Any], raw: dict[str, Any]) -> dict[str, Any]:
    box_office = raw.get("summary", {}).get("box_office", 0)
    customers = snapshot["customer_count"] or 0
    return {
        "date": snapshot["date"],
        "revenue": snapshot["revenue"],
        "box_office": box_office,
        "concession_revenue": round(_filtered_concession_revenue(raw), 2),
        "customer_count": customers,
        "screenings": snapshot["orders"],
        "occupancy_rate": snapshot["usage_rate"],
        "avg_order_value": round(box_office / customers, 2) if customers else 0,
        "last_import_time": snapshot["created_at"],
    }


def _empty_overview(
    status: str,
    message: str,
    target_date: str | None = None,
    last_import_time: str | None = None,
) -> dict[str, Any]:
    return {
        "status": status,
        "data_source": DISPLAY_DATA_SOURCE,
        "date": target_date,
        "revenue": 0,
        "box_office": 0,
        "concession_revenue": 0,
        "customer_count": 0,
        "screenings": 0,
        "occupancy_rate": 0,
        "avg_order_value": 0,
        "last_import_time": last_import_time,
        "message": message,
    }


def _latest_success_time(repository: Any) -> str | None:
    return repository.last_successful_sync_time(PLATFORM)


def _selected_snapshot(repository: Any, target_date: str | None = None) -> dict[str, Any] | None:
    if target_date:
        if target_date > date.today().isoformat():
            return None
        return repository.daily_snapshot_for_date(BUSINESS_TYPE, PLATFORM, STORE_ID, target_date)
    return repository.latest_daily_snapshot_for(BUSINESS_TYPE, PLATFORM, STORE_ID, max_date=date.today().isoformat())


def _load_raw(snapshot: dict[str, Any]) -> dict[str, Any]:
    raw_json = snapshot.get("raw_json") or "{}"
    if isinstance(raw_json, dict):
        return raw_json
    try:
        return json.loads(raw_json)
    except json.JSONDecodeError:
        return {}


def _first_number(rows: list[dict[str, Any]], field: str) -> float:
    for row in rows:
        number = _parse_number(row.get(field))
        if number:
            return number
    return 0


def _sum_or_first(rows: list[dict[str, Any]], field: str, prefer_sum: bool = False) -> float:
    numbers = [_parse_number(row.get(field)) for row in rows]
    numbers = [number for number in numbers if number]
    if not numbers:
        return 0
    return round(sum(numbers), 2) if prefer_sum else numbers[0]


def _first_rate(rows: list[dict[str, Any]], field: str) -> float:
    for row in rows:
        rate = _parse_rate(row.get(field))
        if rate:
            return rate
    return 0


def _parse_number(value: Any) -> float:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_cell(value)
    if not text:
        return 0
    match = re.search(r"-?\d+(?:,\d{3})*(?:\.\d+)?|-?\d+(?:\.\d+)?", text)
    if not match:
        return 0
    return float(match.group(0).replace(",", ""))


def _parse_rate(value: Any) -> float:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        number = float(value)
        return round(number / 100, 4) if number > 1 else round(number, 4)
    text = _clean_cell(value)
    if not text:
        return 0
    number = _parse_number(text)
    return round(number / 100, 4) if "%" in text or number > 1 else round(number, 4)


def _parse_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            import xlrd

            parsed = xlrd.xldate_as_datetime(value, 0)
            if parsed.date() >= date(2000, 1, 1):
                return parsed.date().isoformat()
        except Exception:
            return None
    text = _clean_cell(value)
    if not text:
        return None
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if match:
        year, month, day = map(int, match.groups())
        return date(year, month, day).isoformat()
    return None


def _filter_store_rows(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows_with_store = [row for row in records if _clean_cell(row.get("cinema_name"))]
    if not rows_with_store:
        return []
    return [
        row
        for row in rows_with_store
        if STORE_NAME in _clean_cell(row.get("cinema_name")) or "翡翠城" in _clean_cell(row.get("cinema_name"))
    ]


def _date_range_end_from_filename(filename: str) -> str | None:
    matches = re.findall(r"(20\d{2})[-年](\d{1,2})[-月](\d{1,2})", filename)
    if not matches:
        return None
    year, month, day = map(int, matches[-1])
    return date(year, month, day).isoformat()


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s（）()_\-:/：,，]+", "", _clean_cell(value).lower())


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, time):
        return value.isoformat()
    text = str(value).strip()
    return "" if text.lower() == "none" else text
