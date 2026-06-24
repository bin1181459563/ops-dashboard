"""
影院利润/毛利 + 库存/损耗 Excel 解析服务。
支持三张凤凰云智报表：
  1. 利润毛利报表 (header=4)
  2. 货品进销存汇总报表 (header=4)
  3. 实时库存 (header=0)
"""

import io
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class FinanceImportError(ValueError):
    pass


# ── 娱乐项目排除列表 ──────────────────────────────────────
# 顽小游(PS5/Switch)、小铁台球、顽麻社(棋牌)、娱乐(库存分类)、ZWHRWH(股东商品)
EXCLUDED_CATEGORIES = {"顽小游", "小铁台球", "顽麻社", "娱乐", "ZWHRWH"}


def _is_excluded(item: dict) -> bool:
    """判断商品是否属于被排除的娱乐项目"""
    cat = (item.get("category") or "").strip()
    name = (item.get("item_name") or item.get("item_name") or "").strip()
    # 按大类排除
    if cat in EXCLUDED_CATEGORIES:
        return True
    # 按品名关键词兜底（防止分类不标准的情况）
    for kw in ("顽小游", "小铁台球", "顽麻社", "Ps·Switch", "PS5", "VR", "轰趴"):
        if kw in name:
            return True
    return False


# ── 工具函数 ──────────────────────────────────────────────

def _safe_float(value: Any, default: float = 0.0) -> float:
    """安全转浮点数，支持百分号字符串（如 '82.56%' → 0.8256）"""
    if value is None or value == "--" or value == "":
        return default
    try:
        s = str(value).strip()
        if s.endswith("%"):
            return float(s[:-1]) / 100.0
        return float(s)
    except (ValueError, TypeError):
        return default


def _safe_str(value: Any, default: str = "") -> str:
    """安全转字符串"""
    if value is None:
        return default
    return str(value).strip()


def _generate_batch_id() -> str:
    """生成唯一批次ID"""
    return uuid.uuid4().hex[:12]


def _extract_date_range(ws: Any) -> str:
    """从报表前3行提取日期范围"""
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and "日期" in cell:
                # 匹配 "日期：2026-06-01至2026-06-22"
                m = re.search(r"(\d{4}-\d{2}-\d{2}).*?(\d{4}-\d{2}-\d{2})", cell)
                if m:
                    return f"{m.group(1)}~{m.group(2)}"
    return ""


# ── 利润毛利报表解析 ──────────────────────────────────────

def parse_profit_excel(file_bytes: bytes) -> dict[str, Any]:
    """
    解析利润毛利报表。
    返回: {batch_id, date_range, items: [...]}
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    date_range = _extract_date_range(ws)

    # 表头在第5行 (index=4)
    header_row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    header = [_safe_str(c) for c in header_row]

    # 构建列名映射
    col_map: dict[str, int] = {}
    field_mapping = {
        "商品编码": "item_code",
        "商品名称": "item_name",
        "商品类型": "product_type",
        "大类": "category",
        "一级分类": "sub_category",
        "计量单位": "unit",
        "销售数量": "sales_quantity",
        "销售金额（元）": "sales_amount",
        "退货数量": "return_quantity",
        "退货金额（元）": "return_amount",
        "合计数量": "net_quantity",
        "合计金额（元）": "net_amount",
        "平均售价（元）": "avg_price",
        "销售成本金额（元）": "cost_amount",
        "平均成本单价（元）": "avg_cost_price",
        "利润金额（元）": "profit_amount",
        "毛利率": "profit_rate",
    }
    for i, col_name in enumerate(header):
        if col_name in field_mapping:
            col_map[field_mapping[col_name]] = i

    if "item_code" not in col_map or "item_name" not in col_map:
        raise FinanceImportError("利润毛利报表缺少必要列（商品编码、商品名称）")

    items: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        row_list = list(row)
        item_code = _safe_str(row_list[col_map["item_code"]] if col_map["item_code"] < len(row_list) else None)
        item_name = _safe_str(row_list[col_map["item_name"]] if col_map["item_name"] < len(row_list) else None)
        if not item_code and not item_name:
            continue

        item: dict[str, Any] = {
            "item_code": item_code,
            "item_name": item_name,
        }
        for field, col_idx in col_map.items():
            if field in ("item_code", "item_name"):
                continue
            if col_idx < len(row_list):
                val = row_list[col_idx]
                if field in ("product_type", "category", "sub_category", "unit"):
                    item[field] = _safe_str(val)
                else:
                    item[field] = _safe_float(val)
            else:
                item[field] = "" if field in ("product_type", "category", "sub_category", "unit") else 0.0
        # 兜底：如果毛利率为 0 但有净额和利润，从 profit_amount / net_amount 计算
        if item.get("profit_rate", 0) == 0:
            net = item.get("net_amount", 0)
            if net > 0:
                item["profit_rate"] = round(item.get("profit_amount", 0) / net, 4)
        items.append(item)

    wb.close()

    if not items:
        raise FinanceImportError("利润毛利报表未解析到数据行")

    batch_id = _generate_batch_id()
    return {
        "batch_id": batch_id,
        "date_range": date_range,
        "items": items,
        "file_type": "profit",
        "item_count": len(items),
    }


# ── 进销存汇总报表解析 ──────────────────────────────────

def parse_movement_excel(file_bytes: bytes) -> dict[str, Any]:
    """
    解析货品进销存汇总报表。
    返回: {batch_id, date_range, items: [...]}
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    date_range = _extract_date_range(ws)

    header_row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    header = [_safe_str(c) for c in header_row]

    col_map: dict[str, int] = {}
    field_mapping = {
        "货品编码": "item_code",
        "货品名称": "item_name",
        "大类": "category",
        "一级分类": "sub_category",
        "单位": "unit",
        "期初数量": "opening_qty",
        "期初金额": "opening_amount",
        "进货数量": "purchase_qty",
        "进货金额": "purchase_amount",
        "返货数量": "return_qty",
        "返货金额": "return_amount",
        "调拨入库数量": "transfer_in_qty",
        "调拨入库金额": "transfer_in_amount",
        "调拨出库数量": "transfer_out_qty",
        "出货数量": "outbound_qty",
        "出货金额": "outbound_amount",
        "耗损数量": "loss_qty",
        "耗损金额": "loss_amount",
        "销售数量": "sales_qty",
        "销售成本金额": "sales_cost",
        "盘盈亏数量": "inventory_profit_qty",
        "盘盈亏金额": "inventory_profit_amount",
        "期末数量": "closing_qty",
        "期末金额": "closing_amount",
        "损耗差异比%": "loss_diff_pct",
        "本期移动加权平均成本价": "avg_cost",
    }
    for i, col_name in enumerate(header):
        if col_name in field_mapping:
            col_map[field_mapping[col_name]] = i

    if "item_code" not in col_map or "item_name" not in col_map:
        raise FinanceImportError("进销存汇总报表缺少必要列（货品编码、货品名称）")

    items: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        row_list = list(row)
        item_code = _safe_str(row_list[col_map["item_code"]] if col_map["item_code"] < len(row_list) else None)
        item_name = _safe_str(row_list[col_map["item_name"]] if col_map["item_name"] < len(row_list) else None)
        if not item_code and not item_name:
            continue

        item: dict[str, Any] = {
            "item_code": item_code,
            "item_name": item_name,
        }
        for field, col_idx in col_map.items():
            if field in ("item_code", "item_name"):
                continue
            if col_idx < len(row_list):
                val = row_list[col_idx]
                if field in ("category", "sub_category", "unit"):
                    item[field] = _safe_str(val)
                else:
                    item[field] = _safe_float(val)
            else:
                item[field] = "" if field in ("category", "sub_category", "unit") else 0.0
        items.append(item)

    wb.close()

    if not items:
        raise FinanceImportError("进销存汇总报表未解析到数据行")

    batch_id = _generate_batch_id()
    return {
        "batch_id": batch_id,
        "date_range": date_range,
        "items": items,
        "file_type": "movement",
        "item_count": len(items),
    }


# ── 实时库存解析 ──────────────────────────────────────────

def parse_inventory_excel(file_bytes: bytes) -> dict[str, Any]:
    """
    解析实时库存报表 (header=0)。
    返回: {batch_id, items: [...]}
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header = [_safe_str(c) for c in header_row]

    col_map: dict[str, int] = {}
    field_mapping = {
        "卖品编码": "item_code",
        "卖品名称": "item_name",
        "一级分类": "category",
        "库存": "stock_quantity",
        "含税库存成本": "stock_cost",
        "POS零售价": "pos_price",
    }
    for i, col_name in enumerate(header):
        if col_name in field_mapping:
            col_map[field_mapping[col_name]] = i

    if "item_code" not in col_map or "item_name" not in col_map:
        raise FinanceImportError("实时库存报表缺少必要列（卖品编码、卖品名称）")

    items: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_list = list(row)
        item_code = _safe_str(row_list[col_map["item_code"]] if col_map["item_code"] < len(row_list) else None)
        item_name = _safe_str(row_list[col_map["item_name"]] if col_map["item_name"] < len(row_list) else None)
        if not item_code and not item_name:
            continue

        item: dict[str, Any] = {
            "item_code": item_code,
            "item_name": item_name,
        }
        for field, col_idx in col_map.items():
            if field in ("item_code", "item_name"):
                continue
            if col_idx < len(row_list):
                val = row_list[col_idx]
                if field == "category":
                    item[field] = _safe_str(val)
                else:
                    item[field] = _safe_float(val)
            else:
                item[field] = "" if field == "category" else 0.0
        items.append(item)

    wb.close()

    if not items:
        raise FinanceImportError("实时库存报表未解析到数据行")

    batch_id = _generate_batch_id()
    return {
        "batch_id": batch_id,
        "items": items,
        "file_type": "inventory",
        "item_count": len(items),
    }


# ── 自动识别报表类型 ──────────────────────────────────────

def detect_finance_file_type(file_bytes: bytes, filename: str = "") -> str:
    """
    自动识别报表类型。
    返回: 'profit' | 'movement' | 'inventory' | 'unknown'
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # 检查前3行标题
    title_text = ""
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                title_text += cell

    if "利润" in title_text or "毛利" in title_text:
        wb.close()
        return "profit"
    if "进销存" in title_text:
        wb.close()
        return "movement"

    # 检查表头特征
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_text = " ".join(_safe_str(c) for c in header_row if c)
    if "卖品编码" in header_text or "库存" in header_text:
        wb.close()
        return "inventory"

    # 检查第5行
    if ws.max_row and ws.max_row >= 5:
        header_row5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
        header_text5 = " ".join(_safe_str(c) for c in header_row5 if c)
        if "利润" in header_text5 or "毛利率" in header_text5:
            wb.close()
            return "profit"
        if "期初" in header_text5 or "进销存" in header_text5 or "货品编码" in header_text5:
            wb.close()
            return "movement"

    wb.close()
    return "unknown"


def parse_finance_file(file_bytes: bytes, filename: str = "") -> dict[str, Any]:
    """
    自动识别并解析财务报表文件。
    返回: {batch_id, date_range?, items, file_type, item_count}
    """
    file_type = detect_finance_file_type(file_bytes, filename)

    if file_type == "profit":
        return parse_profit_excel(file_bytes)
    elif file_type == "movement":
        return parse_movement_excel(file_bytes)
    elif file_type == "inventory":
        return parse_inventory_excel(file_bytes)
    else:
        raise FinanceImportError(f"无法识别报表类型：{filename}，请确认是利润毛利/进销存汇总/实时库存报表")


def save_finance_import(repository: Any, parsed: dict[str, Any]) -> int:
    """
    将解析结果写入数据库。
    返回写入行数。
    """
    file_type = parsed["file_type"]
    batch_id = parsed["batch_id"]
    date_range = parsed.get("date_range", "")

    if file_type == "profit":
        return repository.save_cinema_profit_batch(batch_id, date_range, parsed["items"])
    elif file_type == "movement":
        return repository.save_cinema_movement_batch(batch_id, date_range, parsed["items"])
    elif file_type == "inventory":
        return repository.save_cinema_inventory_batch(batch_id, parsed["items"])
    else:
        raise FinanceImportError(f"未知报表类型：{file_type}")


def get_profit_overview(repository: Any) -> dict[str, Any]:
    """获取利润毛利概览数据"""
    summary = repository.get_cinema_profit_summary()
    if not summary:
        return {"status": "not_imported", "message": "请先导入利润毛利报表"}

    items = summary["items"]
    # 排除娱乐项目（顽小游/小铁台球/顽麻社）
    items = [item for item in items if not _is_excluded(item)]
    # 排除Excel合计行（品名或大类为"--"）
    items = [item for item in items if (item.get("item_name") or "").strip() != "--" and (item.get("category") or "").strip() != "--"]
    # 按大类汇总
    category_stats: dict[str, dict[str, float]] = {}
    total_revenue = 0.0
    total_cost = 0.0
    total_profit = 0.0

    for item in items:
        cat = item.get("category") or "未分类"
        if cat not in category_stats:
            category_stats[cat] = {"revenue": 0, "cost": 0, "profit": 0, "quantity": 0, "items": 0}
        category_stats[cat]["revenue"] += item.get("net_amount", 0)
        category_stats[cat]["cost"] += item.get("cost_amount", 0)
        category_stats[cat]["profit"] += item.get("profit_amount", 0)
        category_stats[cat]["quantity"] += item.get("net_quantity", 0)
        category_stats[cat]["items"] += 1
        total_revenue += item.get("net_amount", 0)
        total_cost += item.get("cost_amount", 0)
        total_profit += item.get("profit_amount", 0)

    overall_margin = round(total_profit / total_revenue * 100, 2) if total_revenue > 0 else 0

    categories = sorted(
        [{"category": k, **v, "margin": round(v["profit"] / v["revenue"] * 100, 2) if v["revenue"] > 0 else 0}
         for k, v in category_stats.items()],
        key=lambda x: x["revenue"],
        reverse=True,
    )

    # 按商品类型汇总
    type_stats: dict[str, dict[str, float]] = {}
    for item in items:
        pt = item.get("product_type") or "未分类"
        if pt not in type_stats:
            type_stats[pt] = {"revenue": 0, "cost": 0, "profit": 0, "quantity": 0}
        type_stats[pt]["revenue"] += item.get("net_amount", 0)
        type_stats[pt]["cost"] += item.get("cost_amount", 0)
        type_stats[pt]["profit"] += item.get("profit_amount", 0)
        type_stats[pt]["quantity"] += item.get("net_quantity", 0)

    product_types = sorted(
        [{"product_type": k, **v, "margin": round(v["profit"] / v["revenue"] * 100, 2) if v["revenue"] > 0 else 0}
         for k, v in type_stats.items()],
        key=lambda x: x["revenue"],
        reverse=True,
    )

    # TOP10 利润商品（按利润金额降序）
    top_profit_items = sorted(items, key=lambda x: x.get("profit_amount", 0), reverse=True)[:10]
    # TOP10 亏损商品（按利润率升序，利润率最低的排前面）
    bottom_items = sorted(items, key=lambda x: x.get("profit_rate", 0))[:10]

    return {
        "status": "ok",
        "batch_id": summary["batch_id"],
        "date_range": summary["date_range"],
        "summary": {
            "total_revenue": round(total_revenue, 2),
            "total_cost": round(total_cost, 2),
            "total_profit": round(total_profit, 2),
            "overall_margin": overall_margin,
            "item_count": len(items),
        },
        "categories": categories,
        "product_types": product_types,
        "top_profit_items": top_profit_items,
        "bottom_items": bottom_items,
        "batches": repository.get_cinema_profit_batches(),
    }


def get_inventory_overview(repository: Any) -> dict[str, Any]:
    """获取库存+进销存概览数据"""
    inv_summary = repository.get_cinema_inventory_summary()
    mov_summary = repository.get_cinema_movement_summary()

    if not inv_summary and not mov_summary:
        return {"status": "not_imported", "message": "请先导入实时库存或进销存报表"}

    result: dict[str, Any] = {"status": "ok"}

    # 实时库存数据
    if inv_summary:
        inv_items = [item for item in inv_summary["items"] if not _is_excluded(item)]
        total_stock_cost = sum(item.get("stock_cost", 0) * item.get("stock_quantity", 0) for item in inv_items)
        total_stock_qty = sum(item.get("stock_quantity", 0) for item in inv_items)
        total_pos_value = sum(item.get("pos_price", 0) * item.get("stock_quantity", 0) for item in inv_items)

        # 按一级分类汇总
        inv_category: dict[str, dict[str, float]] = {}
        for item in inv_items:
            cat = item.get("category") or "未分类"
            if cat not in inv_category:
                inv_category[cat] = {"stock_cost": 0, "stock_qty": 0, "pos_value": 0, "items": 0}
            inv_category[cat]["stock_cost"] += item.get("stock_cost", 0) * item.get("stock_quantity", 0)
            inv_category[cat]["stock_qty"] += item.get("stock_quantity", 0)
            inv_category[cat]["pos_value"] += item.get("pos_price", 0) * item.get("stock_quantity", 0)
            inv_category[cat]["items"] += 1

        result["inventory"] = {
            "batch_id": inv_summary["batch_id"],
            "summary": {
                "total_stock_cost": round(total_stock_cost, 2),
                "total_stock_qty": int(total_stock_qty),
                "total_pos_value": round(total_pos_value, 2),
                "item_count": len(inv_items),
                "potential_margin": round((total_pos_value - total_stock_cost) / total_pos_value * 100, 2) if total_pos_value > 0 else 0,
            },
            "categories": sorted(
                [{"category": k, **v} for k, v in inv_category.items()],
                key=lambda x: x["stock_cost"],
                reverse=True,
            ),
            "items": inv_items,
        }

    # 进销存数据
    if mov_summary:
        mov_items = [item for item in mov_summary["items"] if not _is_excluded(item)]
        total_opening = sum(item.get("opening_amount", 0) for item in mov_items)
        total_purchase = sum(item.get("purchase_amount", 0) for item in mov_items)
        total_return = sum(item.get("return_amount", 0) for item in mov_items)
        total_loss = sum(item.get("loss_amount", 0) for item in mov_items)
        total_sales_qty = sum(item.get("sales_qty", 0) for item in mov_items)
        total_sales_cost = sum(item.get("sales_cost", 0) for item in mov_items)
        total_closing = sum(item.get("closing_amount", 0) for item in mov_items)
        total_inv_profit = sum(item.get("inventory_profit_amount", 0) for item in mov_items)

        # 损耗TOP10（损耗数量在Excel中为负数，取绝对值排序）
        loss_items = sorted(
            [item for item in mov_items if item.get("loss_qty", 0) != 0],
            key=lambda x: abs(x.get("loss_amount", 0)),
            reverse=True,
        )[:10]

        # 按大类汇总进销存
        mov_category: dict[str, dict[str, float]] = {}
        for item in mov_items:
            cat = item.get("category") or "未分类"
            if cat not in mov_category:
                mov_category[cat] = {
                    "opening": 0, "purchase": 0, "return": 0,
                    "loss": 0, "sales_qty": 0, "sales_cost": 0, "closing": 0,
                }
            mov_category[cat]["opening"] += item.get("opening_amount", 0)
            mov_category[cat]["purchase"] += item.get("purchase_amount", 0)
            mov_category[cat]["return"] += item.get("return_amount", 0)
            mov_category[cat]["loss"] += item.get("loss_amount", 0)
            mov_category[cat]["sales_qty"] += item.get("sales_qty", 0)
            mov_category[cat]["sales_cost"] += item.get("sales_cost", 0)
            mov_category[cat]["closing"] += item.get("closing_amount", 0)

        result["movement"] = {
            "batch_id": mov_summary["batch_id"],
            "date_range": mov_summary["date_range"],
            "summary": {
                "opening_amount": round(total_opening, 2),
                "purchase_amount": round(total_purchase, 2),
                "return_amount": round(total_return, 2),
                "loss_amount": round(total_loss, 2),
                "sales_qty": round(total_sales_qty, 2),
                "sales_cost": round(total_sales_cost, 2),
                "closing_amount": round(total_closing, 2),
                "inventory_profit_amount": round(total_inv_profit, 2),
                "item_count": len(mov_items),
            },
            "categories": sorted(
                [{"category": k, **v} for k, v in mov_category.items()],
                key=lambda x: x["loss"],
                reverse=True,
            ),
            "loss_items": loss_items,
            "items": mov_items,
        }

    # 批次历史
    result["inventory_batches"] = repository.get_cinema_inventory_batches()
    result["movement_batches"] = repository.get_cinema_movement_batches()

    return result
