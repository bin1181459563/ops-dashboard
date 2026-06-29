"""
库存预警API - 从凤凰云智实时库存页面获取数据并预警
"""
from datetime import date
from typing import Any
from pathlib import Path
import json

from fastapi import APIRouter, Request
from pydantic import BaseModel

router = APIRouter()

# 默认阈值配置
DEFAULT_THRESHOLDS = {
    "爆米花": 50,
    "可乐": 100,
    "矿泉水": 100,
    "薯片": 30,
    "咖啡": 20,
    "奶茶": 20,
    "果汁": 30,
    "冰淇淋": 20,
    "巧克力": 20,
    "饼干": 20,
    "坚果": 20,
    "果冻": 20,
    "酸奶": 30,
    "面包": 20,
    "蛋糕": 15,
    "披萨": 10,
    "汉堡": 10,
    "热狗": 10,
    "鸡米花": 20,
    "薯条": 30,
    "鸡翅": 20,
    "鸡腿": 20,
    "鸡排": 20,
    "鸡块": 20,
    "鸡柳": 20,
}
DEFAULT_THRESHOLD = 20  # 默认阈值

# 配置文件路径
CONFIG_FILE = Path.home() / ".hermes" / "workspace" / "inventory_alert_config.json"


class ThresholdConfig(BaseModel):
    """阈值配置"""
    thresholds: dict[str, int] = {}  # 商品名关键词 -> 阈值
    default_threshold: int = DEFAULT_THRESHOLD
    excluded_products: list[str] = []  # 不需要提醒的产品


def load_config() -> ThresholdConfig:
    """加载配置"""
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, "r") as f:
                data = json.load(f)
                return ThresholdConfig(**data)
        except Exception:
            pass
    return ThresholdConfig(thresholds=DEFAULT_THRESHOLDS)


def save_config(config: ThresholdConfig):
    """保存配置"""
    CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump(config.model_dump(), f, ensure_ascii=False, indent=2)


def get_threshold(item_name: str, config: ThresholdConfig) -> int:
    """获取商品的预警阈值"""
    # 检查是否在排除列表
    for excluded in config.excluded_products:
        if excluded in item_name:
            return -1  # 表示排除
    
    # 精确商品名优先，避免"可乐/薯片/饼干"等泛关键词覆盖单品配置
    if item_name in config.thresholds:
        return config.thresholds[item_name]

    # 再按关键词长度倒序匹配，优先使用更具体的关键词
    for keyword, threshold in sorted(config.thresholds.items(), key=lambda item: len(item[0]), reverse=True):
        if keyword in item_name:
            return threshold
    
    return config.default_threshold


def _read_excel_items(filepath: Path) -> dict[str, dict]:
    """读取Excel中的库存数据，返回 {商品名: {stock, category, pos_price}}"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # 找表头行
    header_row = None
    for row in range(1, 10):
        vals = [str(c.value or "").strip() for c in ws[row]]
        if any("卖品名称" in v or "商品名称" in v for v in vals):
            header_row = row
            break
    
    if not header_row:
        wb.close()
        return {}
    
    # 找列索引
    headers = [str(c.value or "").strip() for c in ws[header_row]]
    name_col = next((i for i, h in enumerate(headers) if "名称" in h), None)
    qty_col = next((i for i, h in enumerate(headers) if "库存" in h and "成本" not in h), None)
    category_col = next((i for i, h in enumerate(headers) if "分类" in h or "大类" in h), None)
    pos_col = next((i for i, h in enumerate(headers) if "POS" in h or "零售" in h), None)
    
    if name_col is None or qty_col is None:
        wb.close()
        return {}
    
    items = {}
    for row in ws.iter_rows(min_row=header_row + 1):
        name = str(row[name_col].value or "").strip()
        if not name:
            continue
        try:
            qty = float(row[qty_col].value or 0)
        except (ValueError, TypeError):
            qty = 0
        category = str(row[category_col].value or "").strip() if category_col is not None else ""
        try:
            pos_price = float(row[pos_col].value or 0) if pos_col is not None else 0
        except (ValueError, TypeError):
            pos_price = 0
        
        items[name] = {"stock": qty, "category": category, "pos_price": pos_price}
    
    wb.close()
    return items


def _read_db_inventory(repository) -> tuple[dict[str, dict], dict[str, dict]]:
    """从数据库读取最新的API采集库存数据，返回 (front_items, warehouse_items)"""
    import sqlite3
    
    conn = sqlite3.connect(str(repository.db_path))
    conn.row_factory = sqlite3.Row
    
    # 查找最新的API采集数据
    cursor = conn.execute("""
        SELECT raw_json FROM daily_snapshots 
        WHERE business_type = 'cinema' AND platform = 'fenghuang'
        ORDER BY date DESC LIMIT 1
    """)
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return {}, {}
    
    raw = json.loads(row["raw_json"])
    inventory_items = raw.get("inventory_items") or []
    
    front_items = {}
    warehouse_items = {}
    
    for item in inventory_items:
        name = item.get("item_name", "")
        if not name:
            continue
        
        qty = item.get("stock_quantity", 0)
        category = item.get("category", "")
        pos_price = item.get("pos_price", 0)
        location = item.get("location", "front")
        
        if location == "front":
            front_items[name] = {"stock": qty, "category": category, "pos_price": pos_price}
        else:
            warehouse_items[name] = {"stock": qty, "category": category, "pos_price": pos_price}
    
    return front_items, warehouse_items


@router.get("/cinema/inventory-alert")
def get_inventory_alert(request: Request):
    """获取库存预警数据"""
    import sqlite3
    
    repository = request.app.state.repository
    db_path = repository.db_path
    
    # 从数据库读取最新的库存数据
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    
    # 查找最新的库存快照
    cursor = conn.execute("""
        SELECT raw_json FROM daily_snapshots 
        WHERE business_type = 'cinema' 
        AND platform = 'fenghuang' 
        AND store_id = 'cinema_feicuicheng'
        ORDER BY date DESC 
        LIMIT 1
    """)
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return {"status": "no_data", "message": "暂无库存数据"}
    
    raw = json.loads(row["raw_json"])
    inventory_items = raw.get("inventory_items") or []
    
    if not inventory_items:
        return {"status": "no_data", "message": "暂无库存数据"}
    
    # 加载配置
    config = load_config()
    
    # 分析库存预警
    low_stock_items = []
    for item in inventory_items:
        item_name = item.get("item_name", "")
        stock_quantity = item.get("stock_quantity", 0)
        
        # 获取阈值
        threshold = get_threshold(item_name, config)
        if threshold == -1:  # 排除的产品
            continue
        
        # 检查是否低于阈值
        if stock_quantity < threshold:
            low_stock_items.append({
                "item_name": item_name,
                "category": item.get("category", ""),
                "stock_quantity": stock_quantity,
                "threshold": threshold,
                "shortage": threshold - stock_quantity,
                "stock_cost": item.get("stock_cost", 0),
                "pos_price": item.get("pos_price", 0),
            })
    
    # 按缺口数量排序
    low_stock_items.sort(key=lambda x: -x["shortage"])
    
    # 统计
    total_items = len(inventory_items)
    alert_items = len(low_stock_items)
    total_shortage = sum(item["shortage"] for item in low_stock_items)
    
    return {
        "status": "ok",
        "date": raw.get("date", ""),
        "summary": {
            "total_items": total_items,
            "alert_items": alert_items,
            "total_shortage": total_shortage,
            "alert_rate": round(alert_items / total_items * 100, 1) if total_items > 0 else 0,
        },
        "items": low_stock_items,
        "config": {
            "thresholds": config.thresholds,
            "default_threshold": config.default_threshold,
            "excluded_products": config.excluded_products,
        },
    }


@router.get("/cinema/inventory-alert/all-items")
def get_all_inventory_items(request: Request):
    """返回全部库存商品（前台+大仓合并），含阈值和排除状态"""
    repository = request.app.state.repository
    
    # 优先从数据库读取API采集的数据
    front_items, warehouse_items = _read_db_inventory(repository)
    data_source = "database"
    
    # 如果数据库没有数据，从Excel读取
    if not front_items:
        data_dir = Path.home() / ".hermes" / "workspace" / "cinema-data"
        front_files = list(data_dir.glob("实时库存_翡翠城店_前台_*.xlsx"))
        
        if not front_files:
            return {"status": "no_data", "message": "未找到前台库存数据"}
        
        latest_front = max(front_files, key=lambda f: f.stat().st_mtime)
        front_items = _read_excel_items(latest_front)
        data_source = "excel"
        
        # 大仓数据
        warehouse_files = list(data_dir.glob("实时库存_翡翠城店_大仓_*.xlsx"))
        if warehouse_files:
            latest_warehouse = max(warehouse_files, key=lambda f: f.stat().st_mtime)
            warehouse_items = _read_excel_items(latest_warehouse)
    
    config = load_config()
    
    # 合并所有商品名（前台+大仓取并集）
    all_names = set(front_items.keys()) | set(warehouse_items.keys())
    
    result = []
    for name in all_names:
        front = front_items.get(name, {"stock": 0, "category": "", "pos_price": 0})
        wh = warehouse_items.get(name, {"stock": 0})
        
        front_stock = front["stock"]
        wh_stock = wh["stock"]
        category = front["category"] or (warehouse_items.get(name) or {}).get("category", "")
        pos_price = front["pos_price"]
        
        threshold = get_threshold(name, config)
        is_excluded = threshold == -1
        eff_threshold = config.default_threshold if is_excluded else threshold
        
        # 前台是否低于阈值
        front_low = not is_excluded and front_stock < eff_threshold
        # 大仓是否无货
        wh_empty = wh_stock <= 0
        
        # 状态标签
        if is_excluded:
            status = "excluded"
        elif front_low and wh_empty:
            status = "critical"  # 前台缺货 + 大仓无货
        elif front_low:
            status = "warning"   # 前台缺货，但大仓有货
        else:
            status = "ok"
        
        result.append({
            "item_name": name,
            "category": category,
            "front_stock": front_stock,
            "wh_stock": wh_stock,
            "threshold": eff_threshold,
            "is_excluded": is_excluded,
            "front_low": front_low,
            "wh_empty": wh_empty,
            "status": status,
            "shortage": 0 if is_excluded else max(0, eff_threshold - front_stock),
            "pos_price": pos_price,
        })
    
    # 排序：critical > warning > ok > excluded，同级按缺口降序
    status_order = {"critical": 0, "warning": 1, "ok": 2, "excluded": 3}
    result.sort(key=lambda x: (status_order.get(x["status"], 9), -x["shortage"]))
    
    return {
        "status": "ok",
        "data_source": data_source,
        "total": len(result),
        "items": result,
        "config": {
            "thresholds": config.thresholds,
            "default_threshold": config.default_threshold,
            "excluded_products": config.excluded_products,
        },
    }


@router.post("/cinema/inventory-alert/config/item")
def update_single_item_config(body: dict):
    """更新单个商品的阈值或排除状态"""
    item_name = body.get("item_name", "")
    action = body.get("action", "")  # "set_threshold" | "exclude" | "include"
    value = body.get("value", 0)
    
    if not item_name or not action:
        return {"status": "error", "message": "缺少参数"}
    
    config = load_config()
    
    if action == "set_threshold":
        config.thresholds[item_name] = int(value)
    elif action == "exclude":
        if item_name not in config.excluded_products:
            config.excluded_products.append(item_name)
    elif action == "include":
        config.excluded_products = [p for p in config.excluded_products if p != item_name]
    else:
        return {"status": "error", "message": f"未知操作: {action}"}
    
    save_config(config)
    return {"status": "ok", "message": f"已更新 {item_name}"}


@router.get("/cinema/inventory-alert/config")
def get_config():
    """获取当前配置"""
    config = load_config()
    return {
        "status": "ok",
        "config": config.model_dump(),
    }
