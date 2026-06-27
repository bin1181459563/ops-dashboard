#!/usr/bin/env python3
"""
导入场次放映明细到 daily_snapshots（带合并逻辑）
不会覆盖已有的卖品/会员数据，只合并场次相关的字段
"""
import sqlite3
import json
from datetime import datetime
from pathlib import Path

import openpyxl

DB_PATH = Path(__file__).parent.parent / "data" / "ops_dashboard.db"
EXCEL_PATH = "/Users/Zhuanz/.hermes-web-ui/upload/default/2a6296c664590982.xlsx"


def merge_snapshot(existing_raw: dict, new_data: dict, date_str: str) -> dict:
    """
    合并场次数据到已有快照。
    保留已有的 concession_revenue / concession_items / member_items，
    只更新 box_office / screenings / customer_count。
    """
    existing_summary = existing_raw.get("summary", {})

    # 新数据的场次字段
    new_box_office = new_data["box_office"]
    new_screenings = new_data["screenings"]
    new_customer = new_data["customer_count"]

    # 保留已有的卖品/会员字段
    best_concession = existing_summary.get("concession_revenue", 0) or 0
    best_member = existing_summary.get("member_consume", 0) or 0
    best_occupancy = existing_summary.get("occupancy_rate", 0) or 0

    # revenue = 票房 + 卖品
    revenue = round(new_box_office + best_concession, 2)
    avg_order_value = round(revenue / new_customer, 2) if new_customer else 0

    merged_raw = {
        **existing_raw,  # 保留已有字段（concession_items, member_items 等）
        "report_type": "screening_detail",
        "date": date_str,
        "summary": {
            "box_office": round(new_box_office, 2),
            "concession_revenue": best_concession,
            "customer_count": new_customer,
            "screenings": new_screenings,
            "occupancy_rate": best_occupancy,
            "member_consume": best_member,
            "revenue": revenue,
            "avg_order_value": avg_order_value,
        },
    }

    return {
        "revenue": revenue,
        "orders": new_screenings,
        "customer_count": new_customer,
        "raw_json": json.dumps(merged_raw, ensure_ascii=False),
    }


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    # 表头在第5行
    headers = [cell.value for cell in ws[5]]
    print(f"表头: {headers[:15]}")

    # 按日期聚合
    daily = {}  # date -> {box_office, screenings, customer_count}

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        date_str = row[6]  # 放映日期
        if not date_str or date_str == '--':
            continue

        date_str = str(date_str)
        screenings = row[8] or 0  # 场次数
        box_office = row[9] or 0  # 票房总金额
        attendance = row[10] or 0  # 观影总人次

        if date_str not in daily:
            daily[date_str] = {"box_office": 0, "screenings": 0, "customer_count": 0}

        daily[date_str]["box_office"] += float(box_office)
        daily[date_str]["screenings"] += int(float(screenings))
        daily[date_str]["customer_count"] += int(float(attendance))

    print(f"解析完成，共 {len(daily)} 天")

    # 连接数据库
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    now = datetime.now().isoformat()

    # 先清理错误的 box_office 报表数据（累计值，不是每日值）
    deleted = conn.execute(
        "DELETE FROM daily_snapshots WHERE business_type='cinema' "
        "AND json_extract(raw_json, '$.report_type') = 'box_office'"
    ).rowcount
    if deleted:
        print(f"清理了 {deleted} 条错误的累计票房报表数据")

    success = 0
    merged_count = 0

    for date_str in sorted(daily.keys()):
        data = daily[date_str]

        # 查询已有数据
        existing = conn.execute(
            "SELECT raw_json FROM daily_snapshots "
            "WHERE business_type='cinema' AND platform='fenghuang' AND store_id='cinema_feicuicheng' AND date=?",
            (date_str,)
        ).fetchone()

        if existing and existing["raw_json"]:
            # 有已有数据 → 合并
            existing_raw = json.loads(existing["raw_json"])
            merged = merge_snapshot(existing_raw, data, date_str)
            merged_count += 1
        else:
            # 无已有数据 → 新建
            revenue = round(data["box_office"], 2)
            raw = {
                "report_type": "screening_detail",
                "date": date_str,
                "summary": {
                    "box_office": round(data["box_office"], 2),
                    "concession_revenue": 0,
                    "customer_count": data["customer_count"],
                    "screenings": data["screenings"],
                    "revenue": revenue,
                    "avg_order_value": round(revenue / data["customer_count"], 2) if data["customer_count"] else 0,
                }
            }
            merged = {
                "revenue": revenue,
                "orders": data["screenings"],
                "customer_count": data["customer_count"],
                "raw_json": json.dumps(raw, ensure_ascii=False),
            }

        conn.execute(
            """
            INSERT INTO daily_snapshots (business_type, platform, store_id, date, revenue, orders, customer_count, raw_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(business_type, platform, store_id, date) DO UPDATE SET
                revenue = excluded.revenue,
                orders = excluded.orders,
                customer_count = excluded.customer_count,
                raw_json = excluded.raw_json,
                created_at = excluded.created_at
            """,
            (
                "cinema", "fenghuang", "cinema_feicuicheng",
                date_str, merged["revenue"], merged["orders"],
                merged["customer_count"], merged["raw_json"], now,
            )
        )
        success += 1

    conn.commit()

    # 验证结果
    stats = conn.execute(
        "SELECT COUNT(*), SUM(revenue), SUM(customer_count) FROM daily_snapshots "
        "WHERE business_type='cinema'"
    ).fetchone()

    concession_count = conn.execute(
        "SELECT COUNT(*) FROM daily_snapshots "
        "WHERE business_type='cinema' AND json_extract(raw_json, '$.summary.concession_revenue') > 0"
    ).fetchone()[0]

    conn.close()

    print(f"\n导入完成: {success} 天（其中 {merged_count} 天合并了已有数据）")
    print(f"数据库总计: {stats[0]} 天, 总收入 ¥{stats[1]:.0f}, 总人次 {stats[2]}")
    print(f"有卖品数据的天数: {concession_count}")

    # 打印前几天和最后几天
    dates = sorted(daily.keys())
    for d in dates[:3]:
        r = daily[d]
        print(f"  {d}: 票房¥{r['box_office']:.0f}, {r['screenings']}场, {r['customer_count']}人")
    print("  ...")
    for d in dates[-3:]:
        r = daily[d]
        print(f"  {d}: 票房¥{r['box_office']:.0f}, {r['screenings']}场, {r['customer_count']}人")


if __name__ == "__main__":
    main()
