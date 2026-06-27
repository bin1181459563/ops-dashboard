#!/usr/bin/env python3
"""导入卖品销售明细到 daily_snapshots（合并模式，不覆盖场次数据）"""
import sqlite3
import json
from datetime import datetime
from pathlib import Path
import openpyxl

DB_PATH = Path(__file__).parent.parent / "data" / "ops_dashboard.db"
EXCEL_PATH = "/Users/Zhuanz/.hermes-web-ui/upload/default/25ded1ea93637385.xlsx"

# 排除的娱乐项（ZWHRWH是股东商品，不排除）
EXCLUDED_CATEGORIES = {"顽小游", "小铁台球", "顽麻社", "娱乐"}
EXCLUDED_KEYWORDS = ["顽小游", "小铁台球", "顽麻社", "Ps·Switch", "PS5", "VR", "轰趴"]


def is_excluded(category: str, item_name: str) -> bool:
    if category in EXCLUDED_CATEGORIES:
        return True
    for kw in EXCLUDED_KEYWORDS:
        if kw in item_name:
            return True
    return False


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    print(f"行数: {ws.max_row}")

    # 按日期聚合卖品收入（排除娱乐项）
    daily = {}       # date -> total revenue
    daily_items = {} # date -> [{category, name, revenue}]
    excluded_count = 0

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        date_str = row[0]  # 销售日期
        if not date_str:
            continue
        date_str = str(date_str)[:10]  # 只取日期部分
        # 验证日期格式
        if not date_str.startswith("20") or len(date_str) != 10:
            continue

        category = str(row[9] or "")  # 卖品大类
        item_name = str(row[15] or "")  # 卖品名称
        quantity = row[17]  # 销售数量
        price = row[20]  # 实际售价（元）

        if is_excluded(category, item_name):
            excluded_count += 1
            continue

        try:
            revenue = float(price) if price else 0
        except (ValueError, TypeError):
            revenue = 0
        try:
            qty = int(quantity) if quantity is not None else 1
        except (ValueError, TypeError):
            qty = 1

        if date_str not in daily:
            daily[date_str] = 0
            daily_items[date_str] = []
        daily[date_str] += revenue
        daily_items[date_str].append({"category": category, "item_name": item_name, "quantity": qty, "revenue": round(revenue, 2)})

    print(f"解析完成: {len(daily)} 天, 排除 {excluded_count} 条娱乐项")

    # 写入数据库（合并模式）
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    now = datetime.now().isoformat()

    success = 0
    merged = 0
    new_insert = 0

    for date_str in sorted(daily.keys()):
        concession_revenue = round(daily[date_str], 2)
        if concession_revenue <= 0:
            continue

        # 查询已有数据
        existing = conn.execute(
            "SELECT raw_json, revenue, orders, customer_count FROM daily_snapshots "
            "WHERE business_type='cinema' AND platform='fenghuang' AND store_id='cinema_feicuicheng' AND date=?",
            (date_str,)
        ).fetchone()

        if existing and existing["raw_json"]:
            # 合并：保留场次数据，加入卖品
            raw = json.loads(existing["raw_json"])
            summary = raw.get("summary", {})
            box_office = summary.get("box_office", 0) or 0
            screenings = summary.get("screenings", 0) or 0
            customers = summary.get("customer_count", 0) or 0

            summary["concession_revenue"] = concession_revenue
            summary["revenue"] = round(box_office + concession_revenue, 2)
            raw["summary"] = summary
            raw["concession_items"] = daily_items[date_str]  # 保留全部明细

            total_revenue = round(box_office + concession_revenue, 2)
            conn.execute(
                "UPDATE daily_snapshots SET revenue=?, raw_json=?, created_at=? "
                "WHERE business_type='cinema' AND platform='fenghuang' AND store_id='cinema_feicuicheng' AND date=?",
                (total_revenue, json.dumps(raw, ensure_ascii=False), now, date_str)
            )
            merged += 1
        else:
            # 新建（只有卖品，没有场次）
            raw = {
                "report_type": "concession_detail",
                "date": date_str,
                "summary": {
                    "box_office": 0,
                    "concession_revenue": concession_revenue,
                    "customer_count": 0,
                    "screenings": 0,
                    "revenue": concession_revenue,
                },
                "concession_items": daily_items[date_str],
            }
            conn.execute(
                "INSERT INTO daily_snapshots (business_type, platform, store_id, date, revenue, orders, customer_count, raw_json, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("cinema", "fenghuang", "cinema_feicuicheng", date_str, concession_revenue, 0, 0,
                 json.dumps(raw, ensure_ascii=False), now)
            )
            new_insert += 1
        success += 1

    conn.commit()

    # 验证
    stats = conn.execute(
        "SELECT COUNT(*), ROUND(SUM(revenue),0), SUM(customer_count) FROM daily_snapshots WHERE business_type='cinema'"
    ).fetchone()
    concession_days = conn.execute(
        "SELECT COUNT(*), ROUND(SUM(json_extract(raw_json, '$.summary.concession_revenue')),0) "
        "FROM daily_snapshots WHERE business_type='cinema' AND json_extract(raw_json, '$.summary.concession_revenue') > 0"
    ).fetchone()

    conn.close()

    print(f"\n导入完成: {success} 天（合并 {merged} + 新建 {new_insert}）")
    print(f"数据库总计: {stats[0]} 天, 总收入 ¥{stats[1]}, 总人次 {stats[2]}")
    print(f"有卖品数据: {concession_days[0]} 天, 卖品总额 ¥{concession_days[1]}")

    # 打印样例
    dates = sorted(daily.keys())
    for d in dates[:3]:
        print(f"  {d}: 卖品 ¥{daily[d]:.0f}")
    print("  ...")
    for d in dates[-3:]:
        print(f"  {d}: 卖品 ¥{daily[d]:.0f}")


if __name__ == "__main__":
    main()
